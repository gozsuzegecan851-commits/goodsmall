import io
from .folder_link_service import (
    get_or_create_folder_link_config,
    folder_link_config_to_dict,
    update_folder_link_config_from_payload,
    get_folder_link_runtime_config,
)
from .folder_link_check_service import check_folder_link_config
import json
from datetime import datetime
from typing import Any
from pathlib import Path

import httpx
from fastapi import APIRouter, Body, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from sqlalchemy import func, or_
from sqlalchemy.orm import Session
from .deps import (
    admin_login_configured,
    admin_login_credentials_match,
    admin_request_authorized,
    clear_admin_session_cookie,
    get_current_admin_profile,
    get_current_admin_username,
    get_db,
    is_masked_secret,
    mask_secret,
    set_admin_session_cookie,
)
from . import schemas
from .models import (
    BotConfig, ProductCategory, Product, ProductSku, Order, OrderItem, PaymentAddress, PaymentOrder,
    Shipment, ShipmentTrace, ShipmentImportBatch, ShipmentImportError, BotRuntimeState,
    Supplier, ProductSupplierMap, OrderFulfillment, CustomerSession, ProductImportBatch, ProductImportError, ChatKeywordBlock, AdminUser, AnnouncementConfig, AnnouncementReceipt, GlobalFolderLinkConfig
)
from .services.shipment_import_service import import_shipments
from .services.product_import_service import (
    build_template_workbook as build_product_import_template_workbook,
    preview_product_import,
    import_product_file,
    list_product_import_batches,
    list_product_import_errors,
    build_error_workbook as build_product_import_error_workbook,
    delete_product_import_batch,
    retry_product_import_batch,
)
from .services.payment_finalize_service import (
    get_order_and_latest_payment_for_update,
    simulate_payment_finalize_enabled,
    simulate_payment_success,
)
from .services.shipment_export_service import build_shipments_workbook
from .config import settings
from .announcement_media_service import (
    build_media_cache_from_items,
    is_supported_video_filename,
    merge_media_cache,
    normalize_announcement_video,
    pick_album_send_items,
    save_telegram_file_ids,
)
from .folder_link_service import (
    ensure_global_folder_link_config,
    folder_link_to_dict,
    is_folder_link_enabled_for_bot,
    normalize_bot_types,
    update_folder_link_config_from_payload,
    validate_folder_link_url,
)
from .folder_link_check_service import check_folder_link_status, should_recheck
from .jobs.logistics_sync import sync_one_shipment
from .services.order_service import (
    resolve_order_supplier,
    upsert_order_fulfillment,
    mark_order_paid_state,
    mark_order_shipped_state,
    mark_order_completed_state,
    mark_order_cancelled_state,
)
from .services.supplier_api_service import build_supplier_payload, push_order_to_supplier, pull_supplier_status
from .services.chat_service import (
    get_chat_overview, list_sessions, get_session_detail, mark_session_read, close_session, reopen_session, send_session_reply, record_customer_event, list_keyword_blocks, get_effective_keyword_blocks, save_keyword_block, toggle_keyword_block, delete_keyword_block
)
from .services.admin_user_service import (
    admin_user_to_dict,
    authenticate_admin_user,
    create_admin_user,
    delete_admin_user,
    get_admin_user_by_username,
    list_admin_users,
    require_superadmin_user,
    set_admin_user_password,
    touch_admin_login,
    update_admin_user,
)
from .services.data_center_service import (
    get_overview as get_data_center_overview,
    get_supplier_board as get_data_center_supplier_board,
    get_trend as get_data_center_trend,
    get_category_supplier_board as get_data_center_category_supplier_board,
    get_product_ranking as get_data_center_product_ranking,
    get_funnel as get_data_center_funnel,
    get_alerts_trend as get_data_center_alerts_trend,
    export_supplier_board_xlsx as export_data_center_supplier_board_xlsx,
    export_product_ranking_xlsx as export_data_center_product_ranking_xlsx,
)

router = APIRouter(prefix="/admin", tags=["admin"])
@router.get("/admin/folder-link/config")
def admin_get_folder_link_config(
    request: Request,
    db: Session = Depends(get_db),
):
    _require_admin_session(request, db)
    cfg = get_or_create_folder_link_config(db)
    return {
        "ok": True,
        "config": folder_link_config_to_dict(cfg),
    }


@router.post("/admin/folder-link/config")
async def admin_save_folder_link_config(
    request: Request,
    db: Session = Depends(get_db),
):
    _require_admin_session(request, db)
    payload = await request.json()
    cfg = get_or_create_folder_link_config(db)
    cfg = update_folder_link_config_from_payload(db, cfg, payload)
    db.commit()
    db.refresh(cfg)
    return {
        "ok": True,
        "config": folder_link_config_to_dict(cfg),
    }


@router.post("/admin/folder-link/check")
def admin_check_folder_link_config(
    request: Request,
    db: Session = Depends(get_db),
):
    _require_admin_session(request, db)
    cfg = get_or_create_folder_link_config(db)
    result = check_folder_link_config(db, cfg)
    db.commit()
    db.refresh(cfg)
    return {
        "ok": True,
        "result": result,
        "config": folder_link_config_to_dict(cfg),
    }


@router.get("/admin/folder-link/runtime")
def admin_get_folder_link_runtime(
    bot_type: str = "",
    bot_code: str = "",
    db: Session = Depends(get_db),
):
    runtime_cfg = get_folder_link_runtime_config(
        db,
        bot_type=bot_type or "",
        bot_code=bot_code or "",
    )
    return {
        "ok": True,
        "config": runtime_cfg,
    }
UPLOAD_DIR = Path(__file__).resolve().parent / "static" / "uploads"
PROFILE_AUTO_SYNC_STATUS = {
    "enabled": bool(settings.bot_profile_auto_sync_enabled),
    "interval_seconds": int(settings.bot_profile_auto_sync_seconds or 3600),
    "scope": str(settings.bot_profile_auto_sync_scope or "enabled"),
    "bot_type": str(settings.bot_profile_auto_sync_bot_type or "all"),
    "last_started_at": "",
    "last_finished_at": "",
    "last_trigger": "",
    "last_total": 0,
    "last_success_count": 0,
    "last_failed_count": 0,
    "last_first_error": "",
}


def _set_profile_auto_sync_status(**kwargs):
    for key, value in kwargs.items():
        PROFILE_AUTO_SYNC_STATUS[key] = value


def get_profile_auto_sync_status() -> dict:
    return dict(PROFILE_AUTO_SYNC_STATUS)

UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

def _telegram_api_url(token: str, method: str) -> str:
    return f"https://api.telegram.org/bot{token}/{method}"


def _telegram_json(token: str, method: str, payload: dict | None = None):
    with httpx.Client(timeout=40.0) as client:
        res = client.post(_telegram_api_url(token, method), json=payload or {})
    res.raise_for_status()
    data = res.json()
    if not data.get("ok"):
        raise RuntimeError(data.get("description") or f"Telegram {method} 调用失败")
    return data.get("result")


def _telegram_multipart(token: str, method: str, data: dict, files: dict):
    with httpx.Client(timeout=80.0) as client:
        res = client.post(_telegram_api_url(token, method), data=data, files=files)
    res.raise_for_status()
    payload = res.json()
    if not payload.get("ok"):
        raise RuntimeError(payload.get("description") or f"Telegram {method} 调用失败")
    return payload.get("result")


def _internal_media_url(url: str) -> str:
    raw = str(url or "").strip()
    if not raw:
        return ""
    if raw.startswith("/"):
        return f"http://backend:8000{raw}"
    raw = raw.replace("http://localhost:8002", "http://backend:8000", 1)
    raw = raw.replace("http://127.0.0.1:8002", "http://backend:8000", 1)
    raw = raw.replace("http://localhost:8001", "http://backend:8000", 1)
    raw = raw.replace("http://127.0.0.1:8001", "http://backend:8000", 1)
    return raw


def _clear_announcement_media_cache_json(raw: str, sorts: list[int] | None = None) -> tuple[str, list[int]]:
    try:
        rows = json.loads(raw or '[]')
        if not isinstance(rows, list):
            rows = []
    except Exception:
        rows = []
    target_sorts = {int(x) for x in (sorts or []) if str(x).strip().isdigit()}
    clear_all = not target_sorts
    cleared = []
    new_rows = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        current_sort = int(row.get('sort') or 0)
        new_row = dict(row)
        if clear_all or current_sort in target_sorts:
            new_row['telegram_file_id'] = ''
            new_row['telegram_unique_id'] = ''
            new_row['error'] = ''
            new_row['updated_at'] = datetime.utcnow().isoformat()
            cleared.append(current_sort)
        new_rows.append(new_row)
    return json.dumps(new_rows, ensure_ascii=False), sorted(x for x in cleared if x > 0)


def _send_telegram_media(bot_token: str, chat_id: str, media_type: str, media_url: str, text: str):
    text = str(text or "").strip()
    media_type = str(media_type or "none").strip().lower()
    media_url = str(media_url or "").strip()

    if media_type == "video" and media_url:
        internal_url = _internal_media_url(media_url)
        print(f"[announcement] sendVideo start chat_id={chat_id} src={media_url} internal={internal_url}")
        with httpx.Client(timeout=120.0) as client:
            res = client.get(internal_url)
            res.raise_for_status()
            filename = Path(internal_url.split("?", 1)[0]).name or "announcement_video.mp4"
            mime = res.headers.get("content-type") or "video/mp4"
            file_size = len(res.content or b"")
            print(f"[announcement] sendVideo fetched ok chat_id={chat_id} file={filename} size={file_size} mime={mime}")
            data = {"chat_id": str(chat_id)}
            if text:
                data["caption"] = text[:1024]
            result = _telegram_multipart(
                bot_token,
                "sendVideo",
                data=data,
                files={"video": (filename, res.content, mime)},
            )
            print(f"[announcement] sendVideo success chat_id={chat_id} file={filename}")
            return result

    print(f"[announcement] sendMessage fallback chat_id={chat_id} text_len={len(text)}")
    return _telegram_json(bot_token, "sendMessage", {"chat_id": str(chat_id), "text": text or "商城公告"})


def _avatar_url_to_path(url: str) -> Path | None:
    raw = str(url or "").strip()
    if not raw:
        return None
    candidate = Path(raw)
    if candidate.exists():
        return candidate
    marker = "/static/uploads/"
    if marker in raw:
        filename = raw.split(marker, 1)[1].split("?", 1)[0].split("#", 1)[0]
        if filename:
            local_path = UPLOAD_DIR / Path(filename).name
            if local_path.exists():
                return local_path
    return None


def _sync_bot_profile(item: BotConfig) -> dict:
    token = str(item.bot_token or "").strip()
    if not token:
        raise RuntimeError("Bot Token 为空，无法同步资料")
    result = {"ok": True, "steps": []}
    me = _telegram_json(token, "getMe", {})
    item.telegram_username = str(me.get("username") or item.telegram_username or "")
    _telegram_json(token, "setMyName", {"name": str(item.bot_name or "")[:64]})
    result["steps"].append("name")
    _telegram_json(token, "setMyShortDescription", {"short_description": str(item.bot_short_description or "")[:120]})
    result["steps"].append("short_description")
    _telegram_json(token, "setMyDescription", {"description": str(item.bot_description or "")[:512]})
    result["steps"].append("description")
    avatar_path = _avatar_url_to_path(item.avatar_image or "")
    if avatar_path:
        suffix = avatar_path.suffix.lower()
        if suffix not in {".jpg", ".jpeg"}:
            raise RuntimeError("机器人头像同步到 Telegram 仅支持 JPG/JPEG，请上传 JPG 图片")
        with avatar_path.open("rb") as fp:
            _telegram_multipart(
                token,
                "setMyProfilePhoto",
                {"photo": json.dumps({"type": "static", "photo": "attach://botphoto"}, ensure_ascii=False)},
                {"botphoto": (avatar_path.name, fp, "image/jpeg")},
            )
        result["steps"].append("avatar")
    elif str(item.avatar_image or "").strip() == "":
        _telegram_json(token, "removeMyProfilePhoto", {})
        result["steps"].append("avatar_removed")
    item.last_profile_sync_at = datetime.utcnow()
    item.last_profile_sync_error = ""
    return result


def normalize_admin_payment_status(order: Order, payment: PaymentOrder | None) -> str:
    if not payment:
        return "pending"
    if payment.confirm_status in {"confirmed", "paid", "success"}:
        return "confirmed"
    if order.pay_status == "paid":
        return "confirmed"
    return payment.confirm_status or "pending"

def admin_payment_status_text(status: str) -> str:
    return {"pending": "待支付", "confirmed": "已确认", "expired": "已过期", "failed": "失败"}.get(status, status or "待支付")


def _masked_or_keep(raw_value: str, stored_value: str) -> str:
    raw = str(raw_value or "").strip()
    stored = str(stored_value or "").strip()
    if not raw:
        return stored
    if stored and is_masked_secret(raw, stored):
        return stored
    return raw


def _mask_bot_token(value: str) -> str:
    return mask_secret(value)


def _mask_supplier_secret(value: str) -> str:
    return mask_secret(value)


def _ensure_can_complete(order: Order) -> None:
    if (order.delivery_status or 'not_shipped') not in {'shipped', 'signed'}:
        raise HTTPException(status_code=400, detail='订单未发货，不能直接完成')


LOGISTICS_ALERT_EXCEPTION_KEYWORDS = ["退回", "拒收", "异常", "问题件", "丢失", "破损", "problem", "exception", "reject", "returned"]
LOGISTICS_ALERT_LEVEL_RANK = {"yellow": 1, "orange": 2, "red": 3}



def _require_superadmin_request(request: Request, db: Session) -> AdminUser:
    username = get_current_admin_username(request)
    try:
        return require_superadmin_user(db, username)
    except ValueError as e:
        raise HTTPException(status_code=403, detail=str(e))


def _alert_age_hours(value: datetime | None) -> int:
    if not value:
        return 0
    return max(0, int((datetime.utcnow() - value).total_seconds() // 3600))


def _order_full_address(order: Order) -> str:
    parts = [order.province or '', order.city or '', order.district or '', order.address_detail or '']
    return ''.join([str(x).strip() for x in parts if str(x or '').strip()])


def _build_logistics_alert_payload(db: Session, supplier_code: str | None = None, level: str | None = None, limit: int = 100):
    level = (level or '').strip().lower()
    supplier_code = (supplier_code or '').strip()
    rows: list[dict] = []

    supplier_rows = db.query(Supplier).order_by(Supplier.supplier_code.asc()).all()
    supplier_info_map = {
        (s.supplier_code or '').strip(): {
            'supplier_name': s.supplier_name or '',
            'contact_name': s.contact_name or '',
            'contact_phone': s.contact_phone or '',
            'contact_tg': s.contact_tg or '',
        }
        for s in supplier_rows
        if (s.supplier_code or '').strip()
    }

    fulfillment_rows = (
        db.query(OrderFulfillment, Supplier)
        .join(Supplier, Supplier.id == OrderFulfillment.supplier_id)
        .order_by(OrderFulfillment.id.desc())
        .all()
    )
    fulfillment_info_map: dict[int, dict] = {}
    for fulfillment, supplier in fulfillment_rows:
        if fulfillment.order_id in fulfillment_info_map:
            continue
        fulfillment_info_map[fulfillment.order_id] = {
            'supplier_code': supplier.supplier_code or '',
            'supplier_name': supplier.supplier_name or '',
            'contact_name': supplier.contact_name or '',
            'contact_phone': supplier.contact_phone or '',
            'contact_tg': supplier.contact_tg or '',
            'supplier_order_no': fulfillment.supplier_order_no or '',
        }

    order_items_cache: dict[int, str] = {}

    def order_product_summary(order_id: int) -> str:
        if order_id not in order_items_cache:
            items = db.query(OrderItem).filter(OrderItem.order_id == order_id).order_by(OrderItem.id.asc()).all()
            if items:
                parts = []
                for item in items:
                    qty = int(item.qty or 0)
                    name = (item.product_name or item.sku_code or '').strip()
                    if not name:
                        continue
                    parts.append(f"{name} x{qty}" if qty else name)
                order_items_cache[order_id] = '；'.join(parts[:8])
            else:
                order_items_cache[order_id] = ''
        return order_items_cache[order_id]

    def push_row(*, alert_type: str, alert_name: str, alert_level: str, order: Order, age_hours: int, alert_text: str, shipment: Shipment | None = None, last_time: datetime | None = None):
        fulfillment_info = fulfillment_info_map.get(order.id, {})
        effective_supplier_code = (order.supplier_code or fulfillment_info.get('supplier_code') or '').strip()
        supplier_meta = supplier_info_map.get(effective_supplier_code, {})
        effective_supplier_name = (fulfillment_info.get('supplier_name') or supplier_meta.get('supplier_name') or '').strip()
        contact_name = (fulfillment_info.get('contact_name') or supplier_meta.get('contact_name') or '').strip()
        contact_phone = (fulfillment_info.get('contact_phone') or supplier_meta.get('contact_phone') or '').strip()
        contact_tg = (fulfillment_info.get('contact_tg') or supplier_meta.get('contact_tg') or '').strip()
        supplier_order_no = (fulfillment_info.get('supplier_order_no') or '').strip()

        if supplier_code and effective_supplier_code != supplier_code:
            return
        if level and level != 'all' and alert_level != level:
            return
        rows.append({
            'alert_type': alert_type,
            'alert_name': alert_name,
            'alert_level': alert_level,
            'order_id': order.id,
            'order_no': order.order_no,
            'customer_name': order.customer_name or '',
            'customer_phone': order.customer_phone or '',
            'supplier_code': effective_supplier_code,
            'supplier_name': effective_supplier_name,
            'supplier_contact_name': contact_name,
            'supplier_contact_phone': contact_phone,
            'supplier_contact_tg': contact_tg,
            'supplier_order_no': supplier_order_no,
            'product_summary': order_product_summary(order.id),
            'full_address': _order_full_address(order),
            'delivery_status': order.delivery_status or 'not_shipped',
            'shipment_id': shipment.id if shipment else None,
            'tracking_no': shipment.tracking_no if shipment else (order.tracking_no or ''),
            'courier_company': shipment.courier_company if shipment else (order.courier_company or ''),
            'ship_status': shipment.ship_status if shipment else (order.delivery_status or 'not_shipped'),
            'sync_status': getattr(shipment, 'sync_status', '') if shipment else '',
            'age_hours': int(age_hours or 0),
            'alert_text': alert_text,
            'last_time': last_time.isoformat() if last_time else '',
            'updated_at': (shipment.updated_at.isoformat() if shipment and shipment.updated_at else (order.updated_at.isoformat() if order.updated_at else '')),
        })

    pending_ship_q = db.query(Order).filter(Order.pay_status == 'paid', Order.delivery_status != 'shipped', Order.delivery_status != 'signed')
    for order in pending_ship_q.order_by(Order.paid_at.desc().nullslast(), Order.id.desc()).all():
        base_time = order.paid_at or order.created_at
        age_hours = _alert_age_hours(base_time)
        alert_level = ''
        if age_hours >= 24:
            alert_level = 'red'
        elif age_hours >= 12:
            alert_level = 'orange'
        elif age_hours >= 6:
            alert_level = 'yellow'
        if alert_level:
            push_row(
                alert_type='not_shipped_timeout',
                alert_name='未发货超时',
                alert_level=alert_level,
                order=order,
                age_hours=age_hours,
                alert_text=f'订单已支付 {age_hours} 小时，仍未进入已发货状态',
                last_time=base_time,
            )

    shipment_q = db.query(Shipment, Order).join(Order, Order.id == Shipment.order_id)
    for shipment, order in shipment_q.order_by(Shipment.id.desc()).all():
        ship_status = (shipment.ship_status or '').lower()
        sync_status = (getattr(shipment, 'sync_status', '') or '').lower()
        sync_error = (getattr(shipment, 'sync_error', '') or '').strip()
        last_trace_text = (shipment.last_trace_text or '').strip()

        if ship_status != 'signed' and not shipment.last_trace_time:
            base_time = shipment.created_at or order.shipped_at or order.updated_at
            age_hours = _alert_age_hours(base_time)
            if age_hours >= 24:
                push_row(
                    alert_type='no_first_trace',
                    alert_name='首轨迹超时',
                    alert_level='yellow',
                    order=order,
                    shipment=shipment,
                    age_hours=age_hours,
                    alert_text=f'发货后 {age_hours} 小时仍无首条物流轨迹',
                    last_time=base_time,
                )

        if ship_status != 'signed' and shipment.last_trace_time:
            age_hours = _alert_age_hours(shipment.last_trace_time)
            alert_level = ''
            if age_hours >= 72:
                alert_level = 'red'
            elif age_hours >= 48:
                alert_level = 'orange'
            if alert_level:
                push_row(
                    alert_type='trace_stagnant',
                    alert_name='轨迹停滞',
                    alert_level=alert_level,
                    order=order,
                    shipment=shipment,
                    age_hours=age_hours,
                    alert_text=f'最新轨迹已 {age_hours} 小时未更新',
                    last_time=shipment.last_trace_time,
                )

        if sync_status == 'error' or sync_error:
            age_hours = _alert_age_hours(getattr(shipment, 'last_sync_at', None) or shipment.updated_at)
            push_row(
                alert_type='sync_error',
                alert_name='同步异常',
                alert_level='orange',
                order=order,
                shipment=shipment,
                age_hours=age_hours,
                alert_text=(sync_error or '物流同步异常，请人工复查')[:240],
                last_time=getattr(shipment, 'last_sync_at', None) or shipment.updated_at,
            )

        combined = f"{ship_status} {last_trace_text}".lower()
        if ship_status in {'returned', 'exception', 'problem', 'rejected'} or any(k in combined for k in LOGISTICS_ALERT_EXCEPTION_KEYWORDS):
            age_hours = _alert_age_hours(shipment.last_trace_time or getattr(shipment, 'last_sync_at', None) or shipment.updated_at)
            push_row(
                alert_type='logistics_exception',
                alert_name='异常件',
                alert_level='red',
                order=order,
                shipment=shipment,
                age_hours=age_hours,
                alert_text=(last_trace_text or '物流状态异常，请人工处理')[:240],
                last_time=shipment.last_trace_time or getattr(shipment, 'last_sync_at', None) or shipment.updated_at,
            )

    rows.sort(key=lambda r: (LOGISTICS_ALERT_LEVEL_RANK.get(r['alert_level'], 0), int(r.get('age_hours') or 0), r.get('updated_at') or ''), reverse=True)
    if limit and limit > 0:
        rows = rows[:limit]

    overview = {
        'total': len(rows),
        'by_level': {'yellow': 0, 'orange': 0, 'red': 0},
        'by_type': {'not_shipped_timeout': 0, 'no_first_trace': 0, 'trace_stagnant': 0, 'sync_error': 0, 'logistics_exception': 0},
    }
    for row in rows:
        if row['alert_level'] in overview['by_level']:
            overview['by_level'][row['alert_level']] += 1
        if row['alert_type'] in overview['by_type']:
            overview['by_type'][row['alert_type']] += 1
        else:
            overview['by_type'][row['alert_type']] = overview['by_type'].get(row['alert_type'], 0) + 1

    return {
        'supplier_code': supplier_code,
        'level': level or 'all',
        'generated_at': datetime.utcnow().isoformat(),
        'overview': overview,
        'rows': rows,
    }


def _normalize_page(page: int | None, default: int = 1) -> int:
    try:
        value = int(page or default)
    except Exception:
        value = default
    return value if value > 0 else default


def _normalize_page_size(page_size: int | None, default: int = 20, minimum: int = 10, maximum: int = 100) -> int:
    try:
        value = int(page_size or default)
    except Exception:
        value = default
    value = max(minimum, value)
    return min(value, maximum)


def _paged_result(rows: list[dict], total: int, page: int, page_size: int, **extra):
    total = max(int(total or 0), 0)
    page = _normalize_page(page, 1)
    page_size = _normalize_page_size(page_size, default=page_size or 20)
    total_pages = max(1, (total + page_size - 1) // page_size) if total else 1
    if page > total_pages:
        page = total_pages
    data = {
        "rows": rows,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "has_prev": page > 1,
        "has_next": page < total_pages,
    }
    data.update(extra)
    return data


def bot_to_dict(item: BotConfig):
    masked_token = _mask_bot_token(item.bot_token)
    return {
        "id": item.id,
        "bot_code": item.bot_code,
        "bot_token": masked_token,
        "bot_token_masked": masked_token,
        "bot_type": item.bot_type,
        "supplier_code": item.supplier_code or "",
        "bot_name": item.bot_name or "",
        "bot_alias": item.bot_alias or "",
        "bot_short_description": item.bot_short_description or "",
        "bot_description": item.bot_description or "",
        "start_welcome_text": item.start_welcome_text or "",
        "avatar_image": item.avatar_image or "",
        "telegram_username": item.telegram_username or "",
        "last_profile_sync_at": item.last_profile_sync_at.isoformat() if item.last_profile_sync_at else "",
        "last_profile_sync_error": item.last_profile_sync_error or "",
        "is_enabled": item.is_enabled,
        "updated_at": item.updated_at.isoformat() if item.updated_at else "",
    }

def category_to_dict(r: ProductCategory):
    return {
        "id": r.id,
        "name": r.name,
        "cover_image": r.cover_image or "",
        "sort_order": int(r.sort_order or 100),
        "is_active": bool(r.is_active),
    }

def product_sku_to_dict(s: ProductSku) -> dict:
    return {
        "id": s.id,
        "sku_code": s.sku_code or "",
        "sku_name": s.sku_name or "",
        "spec_text": s.spec_text or "",
        "price_cny": str(s.price_cny),
        "original_price_cny": str(s.original_price_cny),
        "stock_qty": int(s.stock_qty or 0),
        "weight_gram": int(s.weight_gram or 0),
        "unit_text": s.unit_text or "件",
        "cover_image": s.cover_image or "",
        "is_active": bool(s.is_active),
        "sort_order": int(s.sort_order or 100),
    }

def pick_default_sku(skus: list[ProductSku]) -> ProductSku | None:
    if not skus:
        return None
    active = [s for s in skus if bool(s.is_active)]
    rows = active or skus
    rows = sorted(rows, key=lambda x: (int(x.sort_order or 100), int(x.id or 0)))
    return rows[0] if rows else None

def product_to_dict(r: Product, db: Session | None = None):
    category_name = ""
    if getattr(r, 'category', None) is not None:
        category_name = r.category.name or ""
    elif db is not None and r.category_id:
        cat = db.query(ProductCategory).filter(ProductCategory.id == r.category_id).first()
        category_name = cat.name if cat else ""
    sku_rows = sorted(list(getattr(r, "skus", []) or []), key=lambda x: (int(x.sort_order or 100), int(x.id or 0)))
    return {
        "id": r.id,
        "category_id": r.category_id,
        "category_name": category_name,
        "name": r.name,
        "subtitle": r.subtitle or "",
        "sku_code": r.sku_code or "",
        "cover_image": r.cover_image or "",
        "gallery_images_json": r.gallery_images_json or "[]",
        "price_cny": str(r.price_cny),
        "original_price_cny": str(r.original_price_cny),
        "stock_qty": int(r.stock_qty or 0),
        "weight_gram": int(r.weight_gram or 0),
        "unit_text": r.unit_text or "件",
        "description": r.description or "",
        "detail_html": r.detail_html or "",
        "is_active": bool(r.is_active),
        "sort_order": int(r.sort_order or 100),
        "sku_list": [product_sku_to_dict(s) for s in sku_rows],
    }

def order_to_dict(r: Order):
    return {
        "id": r.id,
        "order_no": r.order_no,
        "telegram_user_id": r.telegram_user_id,
        "customer_name": r.customer_name,
        "customer_phone": r.customer_phone,
        "province": r.province,
        "city": r.city,
        "district": r.district,
        "address_detail": r.address_detail,
        "postal_code": r.postal_code,
        "payable_amount": str(r.payable_amount),
        "payment_method": r.payment_method,
        "supplier_code": r.supplier_code or "",
        "pay_status": r.pay_status,
        "order_status": r.order_status,
        "delivery_status": r.delivery_status,
        "courier_company": r.courier_company,
        "courier_code": r.courier_code,
        "tracking_no": r.tracking_no,
        "seller_remark": r.seller_remark or "",
        "buyer_remark": r.buyer_remark or "",
        "created_at": r.created_at.isoformat() if r.created_at else "",
        "paid_at": r.paid_at.isoformat() if r.paid_at else "",
        "shipped_at": r.shipped_at.isoformat() if r.shipped_at else "",
        "completed_at": r.completed_at.isoformat() if r.completed_at else "",
    }

def payment_address_to_dict(r: PaymentAddress):
    return {
        "id": r.id,
        "address_label": r.address_label or "",
        "address": r.address or "",
        "qr_image": r.qr_image or "",
        "is_active": bool(r.is_active),
        "sort_order": int(r.sort_order or 100),
        "last_used_at": r.last_used_at.isoformat() if r.last_used_at else "",
    }


def supplier_to_dict(r: Supplier):
    api_ready = bool((r.api_base or '').strip() and (r.api_key or '').strip())
    return {
        "id": r.id,
        "supplier_code": r.supplier_code or "",
        "supplier_name": r.supplier_name or "",
        "supplier_type": r.supplier_type or "manual",
        "api_base": r.api_base or "",
        "api_key": r.api_key or "",
        "api_secret": r.api_secret or "",
        "contact_name": r.contact_name or "",
        "contact_phone": r.contact_phone or "",
        "contact_tg": r.contact_tg or "",
        "template_type": r.template_type or "standard",
        "shipping_bot_code": r.shipping_bot_code or "",
        "api_ready": api_ready,
        "is_active": bool(r.is_active),
        "remark": r.remark or "",
    }

def product_supplier_map_to_dict(r: ProductSupplierMap, db: Session):
    product = db.query(Product).filter(Product.id == r.product_id).first()
    supplier = db.query(Supplier).filter(Supplier.id == r.supplier_id).first()
    return {
        "id": r.id,
        "product_id": r.product_id,
        "product_name": product.name if product else "",
        "supplier_id": r.supplier_id,
        "supplier_code": supplier.supplier_code if supplier else "",
        "supplier_name": supplier.supplier_name if supplier else "",
        "supplier_sku": r.supplier_sku or "",
        "priority": int(r.priority or 100),
        "is_default": bool(r.is_default),
        "is_active": bool(r.is_active),
    }

def order_fulfillment_to_dict(r: OrderFulfillment, db: Session):
    order = db.query(Order).filter(Order.id == r.order_id).first()
    supplier = db.query(Supplier).filter(Supplier.id == r.supplier_id).first()
    return {
        "id": r.id,
        "order_id": r.order_id,
        "order_no": order.order_no if order else "",
        "supplier_id": r.supplier_id,
        "supplier_code": supplier.supplier_code if supplier else "",
        "supplier_name": supplier.supplier_name if supplier else "",
        "supplier_order_no": r.supplier_order_no or "",
        "fulfillment_status": r.fulfillment_status or "assigned",
        "assigned_at": r.assigned_at.isoformat() if r.assigned_at else "",
        "accepted_at": r.accepted_at.isoformat() if r.accepted_at else "",
        "shipped_at": r.shipped_at.isoformat() if r.shipped_at else "",
        "sync_status": r.sync_status or "pending",
        "sync_error": r.sync_error or "",
    }

def sync_bot_profiles_batch_once(db: Session, scope: str = "enabled", bot_type: str = "all", ids: list[int] | None = None, trigger: str = "manual") -> dict:
    scope = str(scope or "enabled").strip().lower() or "enabled"
    bot_type = str(bot_type or "all").strip().lower() or "all"
    ids = [int(x) for x in (ids or []) if str(x).strip().isdigit()]
    started_at = datetime.utcnow()
    _set_profile_auto_sync_status(
        last_started_at=started_at.isoformat(),
        last_trigger=str(trigger or "manual"),
        enabled=bool(settings.bot_profile_auto_sync_enabled),
        interval_seconds=int(settings.bot_profile_auto_sync_seconds or 3600),
        scope=str(settings.bot_profile_auto_sync_scope or "enabled"),
        bot_type=str(settings.bot_profile_auto_sync_bot_type or "all"),
    )

    q = db.query(BotConfig)
    if scope == "ids" and ids:
        q = q.filter(BotConfig.id.in_(ids))
    elif scope == "all":
        pass
    else:
        q = q.filter(BotConfig.is_enabled == True)
    if bot_type and bot_type != "all":
        q = q.filter(BotConfig.bot_type == bot_type)
    rows = q.order_by(BotConfig.id.asc()).all()
    results = []
    ok_count = 0
    fail_count = 0
    for item in rows:
        try:
            profile_sync = _sync_bot_profile(item)
            status = "ok"
            error_text = ""
            ok_count += 1
        except Exception as exc:
            item.last_profile_sync_at = datetime.utcnow()
            item.last_profile_sync_error = str(exc)
            profile_sync = {"ok": False, "error": str(exc)}
            status = "failed"
            error_text = str(exc)
            fail_count += 1
        results.append({
            "id": item.id,
            "bot_code": item.bot_code,
            "bot_type": item.bot_type,
            "is_enabled": bool(item.is_enabled),
            "status": status,
            "error": error_text,
            "profile_sync": profile_sync,
        })
    db.commit()
    first_error = ""
    for row in results:
        if row.get("status") == "failed" and row.get("error"):
            first_error = f"{row.get('bot_code') or '-'} / {row.get('error')}"
            break
    finished_at = datetime.utcnow()
    _set_profile_auto_sync_status(
        last_finished_at=finished_at.isoformat(),
        last_total=len(results),
        last_success_count=ok_count,
        last_failed_count=fail_count,
        last_first_error=first_error,
    )
    return {
        "ok": fail_count == 0,
        "scope": scope,
        "bot_type": bot_type,
        "total": len(results),
        "success_count": ok_count,
        "failed_count": fail_count,
        "results": results,
        "trigger": str(trigger or "manual"),
        "started_at": started_at.isoformat(),
        "finished_at": finished_at.isoformat(),
    }


@router.get("/bots/profile-auto-sync-status")
def admin_bot_profile_auto_sync_status():
    return get_profile_auto_sync_status()


@router.post("/bots/profile-auto-sync-run")
def admin_bot_profile_auto_sync_run(payload: dict = Body(default={}), db: Session = Depends(get_db)):
    scope = str((payload or {}).get("scope") or settings.bot_profile_auto_sync_scope or "enabled").strip().lower()
    bot_type = str((payload or {}).get("bot_type") or settings.bot_profile_auto_sync_bot_type or "all").strip().lower()
    return sync_bot_profiles_batch_once(db, scope=scope, bot_type=bot_type, trigger="manual_run")


@router.get("/bots")
def admin_bots(db: Session = Depends(get_db)):
    rows = db.query(BotConfig).order_by(BotConfig.id.asc()).all()
    return [bot_to_dict(r) for r in rows]


@router.post("/bots")
def admin_save_bot(payload: schemas.BotIn, db: Session = Depends(get_db)):
    bot_code = (payload.bot_code or "").strip()
    item = db.query(BotConfig).filter(BotConfig.id == payload.id).first() if payload.id else BotConfig()
    if payload.id and not item:
        raise HTTPException(status_code=404, detail="Bot 不存在")
    bot_token = _masked_or_keep(payload.bot_token, item.bot_token if payload.id else "")
    bot_type = (payload.bot_type or "buyer").strip() or "buyer"
    supplier_code = (payload.supplier_code or "").strip() if bot_type == "shipping" else ""
    if not bot_code:
        raise HTTPException(status_code=400, detail="bot_code 不能为空")
    if not bot_token:
        raise HTTPException(status_code=400, detail="bot_token 不能为空")
    exists_code = db.query(BotConfig).filter(BotConfig.bot_code == bot_code)
    if payload.id:
        exists_code = exists_code.filter(BotConfig.id != payload.id)
    if exists_code.first():
        raise HTTPException(status_code=400, detail="bot_code 已存在")
    exists_token = db.query(BotConfig).filter(BotConfig.bot_token == bot_token)
    if payload.id:
        exists_token = exists_token.filter(BotConfig.id != payload.id)
    if exists_token.first():
        raise HTTPException(status_code=400, detail="bot_token 已被其他机器人使用")
    old_supplier_code = item.supplier_code or ""
    if bot_type == 'shipping':
        if not supplier_code:
            raise HTTPException(status_code=400, detail="供应链机器人必须绑定供应链编码")
        supplier = db.query(Supplier).filter(Supplier.supplier_code == supplier_code).first()
        if not supplier:
            raise HTTPException(status_code=400, detail="绑定的供应链不存在")
        conflict = db.query(BotConfig).filter(BotConfig.bot_type == 'shipping', BotConfig.supplier_code == supplier_code)
        if payload.id:
            conflict = conflict.filter(BotConfig.id != payload.id)
        if conflict.first():
            raise HTTPException(status_code=400, detail="该供应链已绑定其他供应链机器人")
    else:
        supplier = None
    if not payload.id:
        db.add(item)
    item.bot_code = bot_code
    item.bot_token = bot_token
    item.bot_type = bot_type
    item.supplier_code = supplier_code
    item.bot_name = str(payload.bot_name or "").strip()[:64]
    item.bot_alias = str(payload.bot_alias or "").strip()[:128]
    item.bot_short_description = str(payload.bot_short_description or "").strip()[:120]
    item.bot_description = str(payload.bot_description or "").strip()[:512]
    item.start_welcome_text = str(payload.start_welcome_text or "").strip()[:512]
    item.avatar_image = str(payload.avatar_image or "").strip()
    item.is_enabled = bool(payload.is_enabled)
    if payload.id and old_supplier_code and old_supplier_code != supplier_code:
        old_supplier = db.query(Supplier).filter(Supplier.supplier_code == old_supplier_code, Supplier.shipping_bot_code == item.bot_code).first()
        if old_supplier:
            old_supplier.shipping_bot_code = ''
    if supplier:
        supplier.shipping_bot_code = bot_code
    db.commit()
    db.refresh(item)
    sync_result = {"ok": True, "skipped": True}
    try:
        sync_result = _sync_bot_profile(item)
    except Exception as exc:
        item.last_profile_sync_at = datetime.utcnow()
        item.last_profile_sync_error = str(exc)
        sync_result = {"ok": False, "error": str(exc)}
    db.commit()
    db.refresh(item)
    return {"ok": True, "data": bot_to_dict(item), "profile_sync": sync_result}


@router.post("/bots/{bot_id}/sync-profile")
def admin_sync_bot_profile(bot_id: int, db: Session = Depends(get_db)):
    item = db.query(BotConfig).filter(BotConfig.id == bot_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Bot 不存在")
    try:
        result = _sync_bot_profile(item)
    except Exception as exc:
        item.last_profile_sync_at = datetime.utcnow()
        item.last_profile_sync_error = str(exc)
        db.commit()
        db.refresh(item)
        return {"ok": False, "data": bot_to_dict(item), "error": str(exc)}
    db.commit()
    db.refresh(item)
    return {"ok": True, "data": bot_to_dict(item), "profile_sync": result}


@router.post("/bots/sync-profile-batch")
def admin_sync_bot_profile_batch(payload: dict = Body(default={}), db: Session = Depends(get_db)):
    scope = str((payload or {}).get("scope") or "enabled").strip().lower()
    bot_type = str((payload or {}).get("bot_type") or "all").strip().lower()
    ids_raw = (payload or {}).get("ids") or []
    ids: list[int] = []
    if isinstance(ids_raw, list):
        for value in ids_raw:
            try:
                ids.append(int(value))
            except Exception:
                continue
    return sync_bot_profiles_batch_once(db, scope=scope, bot_type=bot_type, ids=ids, trigger="manual_batch")


@router.post("/bots/{bot_id}/enable")
def admin_enable_bot(bot_id: int, db: Session = Depends(get_db)):
    item = db.query(BotConfig).filter(BotConfig.id == bot_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Bot 不存在")
    item.is_enabled = True
    db.commit()
    return {"ok": True}

@router.post("/bots/{bot_id}/disable")
def admin_disable_bot(bot_id: int, db: Session = Depends(get_db)):
    item = db.query(BotConfig).filter(BotConfig.id == bot_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Bot 不存在")
    item.is_enabled = False
    db.commit()
    return {"ok": True}


@router.get("/bots/enabled_tokens")
def admin_enabled_tokens(bot_type: str | None = None, db: Session = Depends(get_db)):
    q = db.query(BotConfig).filter(BotConfig.is_enabled == True)
    if bot_type:
        q = q.filter(BotConfig.bot_type == bot_type)
    rows = q.order_by(BotConfig.id.asc()).all()
    return [{"bot_code": r.bot_code, "bot_token": r.bot_token, "bot_type": r.bot_type, "supplier_code": r.supplier_code or "", "start_welcome_text": r.start_welcome_text or ""} for r in rows]


def runtime_to_dict(r: BotRuntimeState):
    return {
        "id": r.id,
        "bot_code": r.bot_code,
        "bot_type": r.bot_type,
        "run_status": r.run_status,
        "status_text": r.status_text or "",
        "instance_id": r.instance_id or "",
        "last_heartbeat_at": r.last_heartbeat_at.isoformat() if r.last_heartbeat_at else "",
        "started_at": r.started_at.isoformat() if r.started_at else "",
        "stopped_at": r.stopped_at.isoformat() if r.stopped_at else "",
        "last_error": r.last_error or "",
        "updated_at": r.updated_at.isoformat() if r.updated_at else "",
    }

@router.get("/bots/runtime-state")
def admin_bot_runtime_state(db: Session = Depends(get_db)):
    rows = db.query(BotRuntimeState).order_by(BotRuntimeState.bot_code.asc()).all()
    return [runtime_to_dict(r) for r in rows]

@router.post("/bots/runtime/report")
def admin_bot_runtime_report(payload: dict = Body(...), db: Session = Depends(get_db)):
    bot_code = str(payload.get("bot_code") or "").strip()
    if not bot_code:
        raise HTTPException(status_code=400, detail="bot_code 不能为空")
    item = db.query(BotRuntimeState).filter(BotRuntimeState.bot_code == bot_code).first()
    if not item:
        item = BotRuntimeState(bot_code=bot_code)
        db.add(item)
    item.bot_type = str(payload.get("bot_type") or item.bot_type or "buyer")
    item.run_status = str(payload.get("run_status") or item.run_status or "running")
    item.status_text = str(payload.get("status_text") or "")
    item.instance_id = str(payload.get("instance_id") or "")
    if payload.get("heartbeat", True):
        item.last_heartbeat_at = datetime.utcnow()
    if payload.get("run_status") == "starting" and not item.started_at:
        item.started_at = datetime.utcnow()
        item.stopped_at = None
    elif payload.get("run_status") == "running":
        if not item.started_at:
            item.started_at = datetime.utcnow()
        item.stopped_at = None
    elif payload.get("run_status") in {"stopping", "stopped", "disabled"}:
        item.stopped_at = datetime.utcnow()
    last_error = payload.get("last_error")
    if last_error is not None:
        item.last_error = str(last_error)
    db.commit()
    db.refresh(item)
    return {"ok": True, "data": runtime_to_dict(item)}

@router.post("/bots/runtime-report")
def admin_bot_runtime_report_legacy(payload: dict = Body(...), db: Session = Depends(get_db)):
    normalized = dict(payload or {})
    if normalized.get("status") and not normalized.get("run_status"):
        normalized["run_status"] = normalized.get("status")
    if normalized.get("status_note") and not normalized.get("status_text"):
        normalized["status_text"] = normalized.get("status_note")
    if normalized.get("runner_instance_id") and not normalized.get("instance_id"):
        normalized["instance_id"] = normalized.get("runner_instance_id")
    return admin_bot_runtime_report(normalized, db)

@router.get("/categories")
def admin_categories(db: Session = Depends(get_db)):
    rows = db.query(ProductCategory).order_by(ProductCategory.sort_order.asc(), ProductCategory.id.asc()).all()
    return [category_to_dict(r) for r in rows]

@router.post("/categories")
def admin_save_category(payload: schemas.CategoryIn, db: Session = Depends(get_db)):
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="分类名称不能为空")
    item = db.query(ProductCategory).filter(ProductCategory.id == payload.id).first() if payload.id else ProductCategory()
    if payload.id and not item:
        raise HTTPException(status_code=404, detail="分类不存在")
    if not payload.id:
        db.add(item)
    for k, v in payload.model_dump().items():
        if k != "id":
            setattr(item, k, v)
    db.commit()
    db.refresh(item)
    return {"ok": True, "data": category_to_dict(item)}

@router.post("/categories/{category_id}/toggle")
def admin_toggle_category(category_id: int, db: Session = Depends(get_db)):
    item = db.query(ProductCategory).filter(ProductCategory.id == category_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="分类不存在")
    item.is_active = not bool(item.is_active)
    db.commit()
    db.refresh(item)
    row = product_to_dict(item, db=db)
    return {"ok": True, "is_active": item.is_active, "message": ("商品已上架" if item.is_active else "商品已下架"), "item": row, "data": row}

@router.delete("/categories/{category_id}")
def admin_delete_category(category_id: int, db: Session = Depends(get_db)):
    item = db.query(ProductCategory).filter(ProductCategory.id == category_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="分类不存在")
    product_count = db.query(Product).filter(Product.category_id == category_id).count()
    if product_count > 0:
        raise HTTPException(status_code=400, detail="该分类下仍有商品，不能删除")
    db.delete(item)
    db.commit()
    return {"ok": True}

def _product_soft_deleted(item: Product) -> bool:
    return str(item.sku_code or '').startswith('__DELETED__::')

def _mark_product_soft_deleted(item: Product) -> None:
    if _product_soft_deleted(item):
        return
    original_sku = str(item.sku_code or '').strip()
    original_name = str(item.name or '').strip()
    item.sku_code = f'__DELETED__::{item.id or 0}::{original_sku}'
    item.name = original_name or f'已删除商品#{item.id or 0}'
    item.subtitle = ''
    item.is_active = False
    item.sort_order = 999999

@router.get("/products")
def admin_products(db: Session = Depends(get_db)):
    rows = db.query(Product).order_by(Product.sort_order.asc(), Product.id.asc()).all()
    rows = [r for r in rows if not _product_soft_deleted(r)]
    return [product_to_dict(r, db=db) for r in rows]

@router.post("/products")
def admin_save_product(payload: schemas.ProductIn, db: Session = Depends(get_db)):
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="商品名称不能为空")
    if payload.category_id:
        cat = db.query(ProductCategory).filter(ProductCategory.id == payload.category_id).first()
        if not cat:
            raise HTTPException(status_code=400, detail="分类不存在")

    item = db.query(Product).filter(Product.id == payload.id).first() if payload.id else Product()
    if payload.id and not item:
        raise HTTPException(status_code=404, detail="商品不存在")
    if not payload.id:
        db.add(item)

    data = payload.model_dump()
    sku_list = data.pop("sku_list", []) or []

    if not isinstance(sku_list, list) or not sku_list:
        raise HTTPException(status_code=400, detail="至少保留一个 SKU")

    allowed_fields = {
        "category_id", "name", "subtitle", "cover_image", "gallery_images_json",
        "description", "detail_html", "is_active", "sort_order"
    }
    for k, v in data.items():
        if k != "id" and k in allowed_fields:
            setattr(item, k, v)

    db.flush()

    existing = {int(s.id): s for s in (db.query(ProductSku).filter(ProductSku.product_id == item.id).all() or []) if s.id}
    keep_ids = set()
    active_count = 0
    seen_codes = set()

    for row in sku_list:
        row = row or {}
        sku_id = row.get("id")
        sku_code = str(row.get("sku_code") or "").strip()
        sku_name = str(row.get("sku_name") or "").strip()
        spec_text = str(row.get("spec_text") or "").strip()
        price_cny = row.get("price_cny") or 0
        original_price_cny = row.get("original_price_cny") or 0
        stock_qty = int(row.get("stock_qty") or 0)
        weight_gram = int(row.get("weight_gram") or 0)
        unit_text = str(row.get("unit_text") or "件").strip() or "件"
        cover_image = str(row.get("cover_image") or "").strip()
        is_active = bool(row.get("is_active", True))
        sort_order = int(row.get("sort_order") or 100)

        if not sku_name:
            raise HTTPException(status_code=400, detail="SKU 名称不能为空")
        if stock_qty < 0:
            raise HTTPException(status_code=400, detail="SKU 库存不能小于 0")
        if sku_code:
            if sku_code in seen_codes:
                raise HTTPException(status_code=400, detail="同一商品下 SKU 编码不能重复")
            seen_codes.add(sku_code)
        if is_active:
            active_count += 1

        if sku_id and int(sku_id) in existing:
            sku = existing[int(sku_id)]
        else:
            sku = ProductSku(product_id=item.id)
            db.add(sku)

        sku.sku_code = sku_code
        sku.sku_name = sku_name
        sku.spec_text = spec_text
        sku.price_cny = price_cny
        sku.original_price_cny = original_price_cny
        sku.stock_qty = stock_qty
        sku.weight_gram = weight_gram
        sku.unit_text = unit_text
        sku.cover_image = cover_image
        sku.is_active = is_active
        sku.sort_order = sort_order

        db.flush()
        if sku.id:
            keep_ids.add(int(sku.id))

    if active_count <= 0:
        raise HTTPException(status_code=400, detail="至少启用一个 SKU")

    for sku_id, sku in existing.items():
        if sku_id not in keep_ids:
            db.delete(sku)

    db.flush()
    fresh_skus = db.query(ProductSku).filter(ProductSku.product_id == item.id).all()
    default_sku = pick_default_sku(fresh_skus)
    if default_sku:
        item.sku_code = default_sku.sku_code or ""
        item.price_cny = default_sku.price_cny
        item.original_price_cny = default_sku.original_price_cny
        item.stock_qty = default_sku.stock_qty
        item.weight_gram = default_sku.weight_gram
        item.unit_text = default_sku.unit_text or "件"

    db.commit()
    db.refresh(item)
    row = product_to_dict(item, db=db)
    return {"ok": True, "message": "保存成功", "item": row, "data": row}

@router.post("/products/import/preview")
async def admin_product_import_preview(file: UploadFile = File(...), check_images: bool = Form(False), db: Session = Depends(get_db)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="未选择导入文件")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="上传文件为空")
    try:
        return preview_product_import(db, data, check_images=bool(check_images))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/products/import")
async def admin_product_import(file: UploadFile = File(...), operator_name: str = Form("admin_ui"), check_images: bool = Form(False), db: Session = Depends(get_db)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="未选择导入文件")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="上传文件为空")
    try:
        return import_product_file(db, data, file.filename, operator_name=operator_name or 'admin_ui', check_images=bool(check_images))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/products/import-batches")
def admin_product_import_batches(keyword: str | None = None, result_status: str | None = None, db: Session = Depends(get_db)):
    return list_product_import_batches(db, keyword=keyword or '', result_status=result_status or 'all')


@router.get("/products/import-batches/{batch_id}/errors")
def admin_product_import_batch_errors(batch_id: int, db: Session = Depends(get_db)):
    return list_product_import_errors(db, batch_id)


@router.get("/products/import-batches/{batch_id}/errors.xlsx")
def admin_product_import_batch_errors_xlsx(batch_id: int, db: Session = Depends(get_db)):
    rows = list_product_import_errors(db, batch_id)
    data = build_product_import_error_workbook(rows)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="product_import_errors_{batch_id}.xlsx"'},
    )


@router.delete("/products/import-batches/{batch_id}")
def admin_delete_product_import_batch(batch_id: int, db: Session = Depends(get_db)):
    try:
        return delete_product_import_batch(db, batch_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.post("/products/import-batches/{batch_id}/retry")
def admin_retry_product_import_batch(batch_id: int, db: Session = Depends(get_db)):
    try:
        return retry_product_import_batch(db, batch_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/orders")
def admin_orders(
    search: str | None = None,
    pay_status: str | None = None,
    delivery_status: str | None = None,
    supplier_code: str | None = None,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
):
    page = _normalize_page(page, 1)
    page_size = _normalize_page_size(page_size, default=20, minimum=10, maximum=100)
    q = db.query(Order)
    if search:
        s = search.strip()
        if s:
            like = f"%{s}%"
            q = q.filter(
                or_(
                    Order.order_no.ilike(like),
                    Order.customer_name.ilike(like),
                    Order.customer_phone.ilike(like),
                    Order.tracking_no.ilike(like),
                )
            )
    if pay_status and pay_status != 'all':
        q = q.filter(Order.pay_status == pay_status)
    if delivery_status and delivery_status != 'all':
        q = q.filter(Order.delivery_status == delivery_status)
    if supplier_code and supplier_code != 'all':
        q = q.filter(Order.supplier_code == supplier_code)
    total = int(q.order_by(None).count() or 0)
    rows = q.order_by(Order.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return _paged_result(
        [order_to_dict(r) for r in rows],
        total,
        page,
        page_size,
        filters={
            "search": (search or "").strip(),
            "pay_status": pay_status or "all",
            "delivery_status": delivery_status or "all",
            "supplier_code": supplier_code or "all",
        },
    )

@router.get("/orders/{order_id}")
def admin_order_detail(order_id: int, db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    items = db.query(OrderItem).filter(OrderItem.order_id == order.id).all()
    payment = db.query(PaymentOrder).filter(PaymentOrder.order_id == order.id).order_by(PaymentOrder.id.desc()).first()
    shipment = db.query(Shipment).filter(Shipment.order_id == order.id).order_by(Shipment.id.desc()).first()
    fulfillment = db.query(OrderFulfillment).filter(OrderFulfillment.order_id == order.id).order_by(OrderFulfillment.id.desc()).first()
    traces = db.query(ShipmentTrace).filter(ShipmentTrace.shipment_id == shipment.id).order_by(ShipmentTrace.id.desc()).limit(10).all() if shipment else []
    data = order_to_dict(order)
    data.update({
        "items": [{
            "product_name": i.product_name,
            "sku_code": i.sku_code,
            "qty": i.qty,
            "unit_price": str(i.unit_price),
            "subtotal": str(i.subtotal),
        } for i in items],
        "payment": ({
            "pay_method": payment.pay_method,
            "receive_address": payment.receive_address,
            "expected_amount": str(payment.expected_amount),
            "paid_amount": str(payment.paid_amount),
            "txid": payment.txid,
            "confirm_status": normalize_admin_payment_status(order, payment),
            "confirm_status_text": admin_payment_status_text(normalize_admin_payment_status(order, payment)),
            "paid_at": payment.paid_at.isoformat() if payment and payment.paid_at else (order.paid_at.isoformat() if order.paid_at else ""),
            "expired_at": payment.expired_at.isoformat() if payment and payment.expired_at else "",
        } if payment else None),
        "fulfillment": (order_fulfillment_to_dict(fulfillment, db) if fulfillment else None),
        "available_suppliers": [supplier_to_dict(s) for s in db.query(Supplier).filter(Supplier.is_active == True).order_by(Supplier.supplier_code.asc()).all()],
        "shipment": {
            "courier_company": shipment.courier_company,
            "courier_code": shipment.courier_code,
            "tracking_no": shipment.tracking_no,
            "ship_status": shipment.ship_status,
            "last_trace_text": shipment.last_trace_text,
            "last_trace_time": shipment.last_trace_time.isoformat() if shipment.last_trace_time else "",
        } if shipment else None,
        "traces": [{
            "trace_time": t.trace_time.isoformat() if t.trace_time else "",
            "trace_status": t.trace_status,
            "trace_text": t.trace_text,
        } for t in traces],
    })
    return data


@router.post("/orders/{order_id}/payment-refresh")
def admin_refresh_order_payment(order_id: int, db: Session = Depends(get_db)):
    from .services.payment_sync_service import refresh_payment_order_status

    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")

    payment = db.query(PaymentOrder).filter(PaymentOrder.order_id == order.id).order_by(PaymentOrder.id.desc()).first()
    if not payment:
        raise HTTPException(status_code=400, detail="该订单暂无支付单")

    try:
        result = refresh_payment_order_status(db, order, payment)
        db.commit()
        db.refresh(order)
        db.refresh(payment)
    except RuntimeError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

    return {
        "ok": True,
        "result": result,
        "order": order_to_dict(order),
        "payment": {
            "pay_method": payment.pay_method,
            "receive_address": payment.receive_address,
            "expected_amount": str(payment.expected_amount),
            "paid_amount": str(payment.paid_amount),
            "txid": payment.txid,
            "confirm_status": normalize_admin_payment_status(order, payment),
            "confirm_status_text": admin_payment_status_text(normalize_admin_payment_status(order, payment)),
            "paid_at": payment.paid_at.isoformat() if payment.paid_at else (order.paid_at.isoformat() if order.paid_at else ""),
            "expired_at": payment.expired_at.isoformat() if payment.expired_at else "",
        },
    }


@router.post("/orders/{order_id}/simulate-paid")
def admin_simulate_order_paid(order_id: int, request: Request, db: Session = Depends(get_db)):
    profile = get_current_admin_profile(request)
    if not bool(profile.get("is_superadmin")):
        raise HTTPException(status_code=403, detail="仅 superadmin 可执行模拟支付成功")
    if not simulate_payment_finalize_enabled():
        raise HTTPException(status_code=403, detail="当前环境未开启模拟支付成功")

    operator = str(profile.get("username") or get_current_admin_username(request) or "").strip() or "admin"

    try:
        order, payment = get_order_and_latest_payment_for_update(db, order_id)
        result = simulate_payment_success(db, order, payment, operator=operator)
        db.commit()
        db.refresh(order)
        db.refresh(payment)
    except ValueError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    except Exception:
        db.rollback()
        raise

    return {
        "ok": True,
        "result": result,
        "order": order_to_dict(order),
        "payment": {
            "pay_method": payment.pay_method,
            "receive_address": payment.receive_address,
            "expected_amount": str(payment.expected_amount),
            "paid_amount": str(payment.paid_amount),
            "txid": payment.txid,
            "confirm_status": normalize_admin_payment_status(order, payment),
            "confirm_status_text": admin_payment_status_text(normalize_admin_payment_status(order, payment)),
            "paid_at": payment.paid_at.isoformat() if payment.paid_at else (order.paid_at.isoformat() if order.paid_at else ""),
            "expired_at": payment.expired_at.isoformat() if payment.expired_at else "",
        },
    }


@router.post("/orders/{order_id}/mark-paid")
def admin_mark_paid(order_id: int, db: Session = Depends(get_db)):
    item = db.query(Order).filter(Order.id == order_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="订单不存在")
    mark_order_paid_state(db, item, paid_at=datetime.utcnow())
    db.commit()
    return {"ok": True}

@router.post("/orders/{order_id}/ship")
def admin_ship_order(order_id: int, payload: schemas.OrderShipIn, db: Session = Depends(get_db)):
    item = db.query(Order).filter(Order.id == order_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="订单不存在")
    if item.pay_status != 'paid':
        raise HTTPException(status_code=400, detail="订单未支付，不能发货")
    courier_company = payload.courier_company.strip()
    tracking_no = payload.tracking_no.strip()
    if not courier_company or not tracking_no:
        raise HTTPException(status_code=400, detail="快递公司和快递单号不能为空")
    item.courier_company = courier_company
    item.courier_code = payload.courier_code.strip()
    item.tracking_no = tracking_no
    mark_order_shipped_state(db, item, shipped_at=datetime.utcnow())
    shipment = db.query(Shipment).filter(Shipment.order_id == item.id).order_by(Shipment.id.desc()).first()
    if shipment is None:
        shipment = Shipment(order_id=item.id)
        db.add(shipment)
    shipment.courier_company = item.courier_company
    shipment.courier_code = item.courier_code
    shipment.tracking_no = item.tracking_no
    shipment.ship_status = 'shipped'
    shipment.provider_name = settings.logistics_provider or 'kuaidi100'
    shipment.sync_status = 'pending'
    shipment.sync_error = ''
    shipment.last_sync_at = None
    db.commit()
    return {"ok": True}

@router.post("/orders/{order_id}/complete")
def admin_complete_order(order_id: int, db: Session = Depends(get_db)):
    item = db.query(Order).filter(Order.id == order_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="订单不存在")
    _ensure_can_complete(item)
    mark_order_completed_state(db, item, completed_at=datetime.utcnow())
    db.commit()
    return {"ok": True}

@router.post("/orders/{order_id}/cancel")
def admin_cancel_order(order_id: int, db: Session = Depends(get_db)):
    item = db.query(Order).filter(Order.id == order_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="订单不存在")
    if (item.delivery_status or "not_shipped") in {'shipped', 'signed'}:
        raise HTTPException(status_code=400, detail="订单已发货，不能直接取消")
    mark_order_cancelled_state(db, item)
    db.commit()
    return {"ok": True}

@router.post("/orders/{order_id}/remark")
def admin_order_remark(order_id: int, payload: schemas.OrderRemarkIn, db: Session = Depends(get_db)):
    item = db.query(Order).filter(Order.id == order_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="订单不存在")
    item.seller_remark = payload.seller_remark.strip()
    db.commit()
    return {"ok": True}

@router.get("/payment-addresses")
def admin_payment_addresses(db: Session = Depends(get_db)):
    rows = db.query(PaymentAddress).order_by(PaymentAddress.sort_order.asc(), PaymentAddress.id.asc()).all()
    return [payment_address_to_dict(r) for r in rows]


@router.post("/orders/{order_id}/sync-logistics")
def admin_sync_order_logistics(order_id: int, db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    shipment = db.query(Shipment).filter(Shipment.order_id == order.id).order_by(Shipment.id.desc()).first()
    if not shipment:
        raise HTTPException(status_code=400, detail="订单还未发货")
    result = sync_one_shipment(db, shipment)
    db.commit()
    return result

@router.post("/payment-addresses")
def admin_save_payment_address(payload: schemas.PaymentAddressIn, db: Session = Depends(get_db)):
    address = (payload.address or "").strip()
    if not address:
        raise HTTPException(status_code=400, detail="收款地址不能为空")

    exists = db.query(PaymentAddress).filter(PaymentAddress.address == address)
    if payload.id:
        exists = exists.filter(PaymentAddress.id != payload.id)
    if exists.first():
        raise HTTPException(status_code=400, detail="收款地址已存在")

    item = db.query(PaymentAddress).filter(PaymentAddress.id == payload.id).first() if payload.id else PaymentAddress()
    if payload.id and not item:
        raise HTTPException(status_code=404, detail="收款地址不存在")
    if not payload.id:
        db.add(item)

    item.address_label = (payload.address_label or "").strip()
    item.address = address
    item.qr_image = (payload.qr_image or "").strip()
    item.is_active = bool(payload.is_active)
    item.sort_order = int(payload.sort_order or 100)

    db.commit()
    db.refresh(item)
    return {"ok": True, "data": payment_address_to_dict(item)}

@router.post("/payment-addresses/{address_id}/toggle")
def admin_toggle_payment_address(address_id: int, db: Session = Depends(get_db)):
    item = db.query(PaymentAddress).filter(PaymentAddress.id == address_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="收款地址不存在")
    item.is_active = not bool(item.is_active)
    db.commit()
    db.refresh(item)
    row = product_to_dict(item, db=db)
    return {"ok": True, "is_active": item.is_active, "message": ("商品已上架" if item.is_active else "商品已下架"), "item": row, "data": row}

@router.delete("/payment-addresses/{address_id}")
def admin_delete_payment_address(address_id: int, db: Session = Depends(get_db)):
    item = db.query(PaymentAddress).filter(PaymentAddress.id == address_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="收款地址不存在")
    count = db.query(PaymentOrder).filter(PaymentOrder.receive_address == item.address).count()
    if count > 0:
        raise HTTPException(status_code=400, detail="该收款地址已有支付记录，不能删除")
    db.delete(item)
    db.commit()
    return {"ok": True}

@router.post("/payment-addresses/upload-image")
async def admin_upload_payment_qr(file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="未选择文件")
    suffix = Path(file.filename).suffix.lower()
    if suffix not in {".png", ".jpg", ".jpeg", ".webp"}:
        raise HTTPException(status_code=400, detail="仅支持 png/jpg/jpeg/webp 图片")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="上传文件为空")
    filename = f"payqr_{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}{suffix}"
    save_path = UPLOAD_DIR / filename
    save_path.write_bytes(data)
    public_url = f"{settings.backend_public_url}/static/uploads/{filename}"
    return {"ok": True, "url": public_url, "filename": filename}

@router.get("/shipments/pending-summary")
def admin_pending_shipment_summary(supplier_code: str | None = None, row_limit: int = 20, db: Session = Depends(get_db)):
    row_limit = _normalize_page_size(row_limit, default=20, minimum=5, maximum=100)
    q = db.query(Order).filter(Order.pay_status == "paid", Order.delivery_status == "not_shipped")
    if supplier_code:
        q = q.filter(Order.supplier_code == supplier_code)
    total = int(q.order_by(None).count() or 0)
    today = datetime.utcnow().strftime("%Y-%m-%d")
    rows = q.order_by(Order.paid_at.desc().nullslast(), Order.id.desc()).limit(row_limit).all()
    supplier_rows = db.query(func.coalesce(Order.supplier_code, '未分配').label('supplier_key'), func.count(Order.id)).filter(Order.pay_status == "paid", Order.delivery_status == "not_shipped")
    if supplier_code:
        supplier_rows = supplier_rows.filter(Order.supplier_code == supplier_code)
    supplier_rows = supplier_rows.group_by(func.coalesce(Order.supplier_code, '未分配')).all()
    by_supplier = {str(k or '未分配'): int(v or 0) for k, v in supplier_rows}
    return {"pending_shipment_count": total, "supplier_code": supplier_code or "", "rows": [order_to_dict(r) for r in rows], "biz_date": today, "by_supplier": by_supplier, "row_limit": row_limit}

@router.get("/shipments/shipped-summary")
def admin_shipped_summary(supplier_code: str | None = None, biz_date: str | None = None, row_limit: int = 20, db: Session = Depends(get_db)):
    row_limit = _normalize_page_size(row_limit, default=20, minimum=5, maximum=100)
    q = db.query(Order).filter(Order.delivery_status.in_(["shipped", "signed"]))
    date_filters = []
    if biz_date:
        start = datetime.fromisoformat(biz_date + "T00:00:00")
        end = datetime.fromisoformat(biz_date + "T23:59:59.999999")
        date_filters = [Order.shipped_at.isnot(None), Order.shipped_at >= start, Order.shipped_at <= end]
        q = q.filter(*date_filters)
    if supplier_code:
        q = q.filter(Order.supplier_code == supplier_code)
    total = int(q.order_by(None).count() or 0)
    rows = q.order_by(Order.shipped_at.desc().nullslast(), Order.id.desc()).limit(row_limit).all()

    courier_q = db.query(func.coalesce(Order.courier_company, '未填写').label('courier_key'), func.count(Order.id)).filter(Order.delivery_status.in_(["shipped", "signed"]))
    supplier_q = db.query(func.coalesce(Order.supplier_code, '未分配').label('supplier_key'), func.count(Order.id)).filter(Order.delivery_status.in_(["shipped", "signed"]))
    if date_filters:
        courier_q = courier_q.filter(*date_filters)
        supplier_q = supplier_q.filter(*date_filters)
    if supplier_code:
        courier_q = courier_q.filter(Order.supplier_code == supplier_code)
        supplier_q = supplier_q.filter(Order.supplier_code == supplier_code)
    by_courier = {str(k or '未填写'): int(v or 0) for k, v in courier_q.group_by(func.coalesce(Order.courier_company, '未填写')).all()}
    by_supplier = {str(k or '未分配'): int(v or 0) for k, v in supplier_q.group_by(func.coalesce(Order.supplier_code, '未分配')).all()}
    return {"shipped_count": total, "supplier_code": supplier_code or "", "biz_date": biz_date or "", "rows": [order_to_dict(r) for r in rows], "by_courier": by_courier, "by_supplier": by_supplier, "row_limit": row_limit}

@router.get("/shipments/export-pending")
def admin_export_pending_shipments(supplier_code: str | None = None, db: Session = Depends(get_db)):
    filename, buf = build_shipments_workbook(db, mode="pending", supplier_code=supplier_code)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

@router.get("/shipments/export-shipped")
def admin_export_shipped_shipments(supplier_code: str | None = None, biz_date: str | None = None, db: Session = Depends(get_db)):
    filename, buf = build_shipments_workbook(db, mode="shipped", supplier_code=supplier_code, biz_date=biz_date)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

@router.get("/shipments/import-template")
def admin_import_template():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "已发货导入模板"
    ws.append(["订单号", "快递公司", "快递编码", "快递单号", "发货时间", "收件人", "手机号", "备注"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=shipment_import_template.xlsx"})

@router.get("/suppliers/{supplier_id}/template-sample")
def admin_supplier_template_sample(supplier_id: int, mode: str = "pending", db: Session = Depends(get_db)):
    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="供应链不存在")
    export_mode = "shipped" if mode == "shipped" else "pending"
    filename, buf = build_shipments_workbook(db, mode=export_mode, supplier_code=supplier.supplier_code, sample_only=True)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

@router.post("/shipments/import")
async def admin_import_shipments(file: UploadFile = File(...), supplier_code: str = Form(""), db: Session = Depends(get_db)):
    content = await file.read()
    batch = import_shipments(db, content, file.filename or "shipment_import.xlsx", operator_name="admin", supplier_code=supplier_code)
    return {"ok": True, "batch_id": batch.id, "batch_no": batch.batch_no}

@router.get("/shipments/import-batches")
def admin_import_batches(db: Session = Depends(get_db)):
    rows = db.query(ShipmentImportBatch).order_by(ShipmentImportBatch.id.desc()).all()
    return [{"id": r.id, "batch_no": r.batch_no, "file_name": r.file_name, "biz_date": r.biz_date, "supplier_code": r.supplier_code or "", "operator_name": r.operator_name or "", "total_rows": r.total_rows, "success_rows": r.success_rows, "failed_rows": r.failed_rows, "created_at": r.created_at.isoformat() if r.created_at else ""} for r in rows]

@router.get("/shipments/import-batches/{batch_id}/errors")
def admin_import_batch_errors(batch_id: int, db: Session = Depends(get_db)):
    rows = db.query(ShipmentImportError).filter(ShipmentImportError.batch_id == batch_id).order_by(ShipmentImportError.row_no.asc()).all()
    return [{"row_no": r.row_no, "order_no": r.order_no, "tracking_no": r.tracking_no, "error_message": r.error_message, "raw_row_json": r.raw_row_json or ""} for r in rows]


@router.get("/logistics")
def admin_logistics(search: str | None = None, supplier_code: str | None = None, ship_status: str | None = None, db: Session = Depends(get_db)):
    q = db.query(Shipment, Order).join(Order, Order.id == Shipment.order_id)
    if search:
        s = search.strip()
        q = q.filter((Order.order_no.ilike(f"%{s}%")) | (Shipment.tracking_no.ilike(f"%{s}%")) | (Shipment.courier_company.ilike(f"%{s}%")))
    if supplier_code:
        q = q.filter(Order.supplier_code == supplier_code)
    if ship_status and ship_status != 'all':
        q = q.filter(Shipment.ship_status == ship_status)
    rows = q.order_by(Shipment.id.desc()).all()
    return [{
        "shipment_id": s.id,
        "order_id": o.id,
        "order_no": o.order_no,
        "customer_name": o.customer_name,
        "supplier_code": o.supplier_code or "",
        "courier_company": s.courier_company or "",
        "courier_code": s.courier_code or "",
        "tracking_no": s.tracking_no or "",
        "provider_name": s.provider_name or "",
        "ship_status": s.ship_status or "pending",
        "sync_status": getattr(s, 'sync_status', 'pending') or 'pending',
        "sync_error": getattr(s, 'sync_error', '') or '',
        "last_trace_text": s.last_trace_text or "",
        "last_trace_time": s.last_trace_time.isoformat() if s.last_trace_time else "",
        "last_sync_at": getattr(s, 'last_sync_at', None).isoformat() if getattr(s, 'last_sync_at', None) else (s.updated_at.isoformat() if s.updated_at else ""),
    } for s, o in rows]

@router.get("/logistics/alerts")
def admin_logistics_alerts(supplier_code: str | None = None, level: str | None = None, limit: int = 100, db: Session = Depends(get_db)):
    safe_limit = max(1, min(int(limit or 100), 300))
    return _build_logistics_alert_payload(db, supplier_code=supplier_code, level=level, limit=safe_limit)


@router.get("/logistics/alerts/export.xlsx")
def admin_logistics_alerts_export_xlsx(supplier_code: str | None = None, level: str | None = None, db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    payload = _build_logistics_alert_payload(db, supplier_code=supplier_code, level=level, limit=2000)
    rows = payload.get('rows') or []
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '物流预警核对表'

    filter_supplier = (supplier_code or '全部供应链').strip() or '全部供应链'
    filter_level = (level or 'all').strip() or 'all'
    export_time = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')

    ws.append(['物流预警核对表'])
    ws.append([f'导出时间：{export_time}｜供应链：{filter_supplier}｜等级：{filter_level}'])
    ws.append([])
    headers = [
        '预警等级', '预警项', '订单号', '供应链编码', '供应链名称', '供应链联系人', '供应链电话', '供应链TG',
        '供应链单号', '商品信息', '客户姓名', '客户电话', '完整地址', '快递公司', '快递单号', '物流状态', '同步状态',
        '预警说明', '持续时长(小时)', '最近时间'
    ]
    ws.append(headers)

    for row in rows:
        ws.append([
            row.get('alert_level') or '',
            row.get('alert_name') or row.get('alert_type') or '',
            row.get('order_no') or '',
            row.get('supplier_code') or '',
            row.get('supplier_name') or '',
            row.get('supplier_contact_name') or '',
            row.get('supplier_contact_phone') or '',
            row.get('supplier_contact_tg') or '',
            row.get('supplier_order_no') or '',
            row.get('product_summary') or '',
            row.get('customer_name') or '',
            row.get('customer_phone') or '',
            row.get('full_address') or '',
            row.get('courier_company') or '',
            row.get('tracking_no') or '',
            row.get('ship_status') or '',
            row.get('sync_status') or '',
            row.get('alert_text') or '',
            int(row.get('age_hours') or 0),
            row.get('last_time') or row.get('updated_at') or '',
        ])

    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'].font = Font(size=10)
    header_row = 4
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    widths = [10, 12, 24, 14, 18, 14, 16, 16, 18, 30, 12, 16, 36, 14, 20, 12, 12, 32, 12, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row_cells in ws.iter_rows(min_row=5, max_row=ws.max_row):
        for cell in row_cells:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    filename = f'logistics_alerts_{(supplier_code or "all").strip() or "all"}_{(level or "all").strip() or "all"}.xlsx'
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename="{filename}"'})


@router.get("/logistics/api-overview")
def admin_logistics_api_overview(db: Session = Depends(get_db)):
    suppliers = db.query(Supplier).order_by(Supplier.supplier_code.asc()).all()
    return {
        "provider_name": settings.logistics_provider or 'kuaidi100',
        "provider_key_configured": bool(getattr(settings, 'kuaidi100_key', '') or getattr(settings, 'kuaidi100_customer', '')),
        "rows": [{
            "supplier_id": s.id,
            "supplier_code": s.supplier_code or '',
            "supplier_name": s.supplier_name or '',
            "supplier_type": s.supplier_type or 'manual',
            "api_base": s.api_base or '',
            "api_key_configured": bool(s.api_key),
            "template_type": s.template_type or 'standard',
            "shipping_bot_code": s.shipping_bot_code or '',
            "is_active": bool(s.is_active),
        } for s in suppliers],
    }


@router.get("/logistics/{shipment_id}")
def admin_logistics_detail(shipment_id: int, db: Session = Depends(get_db)):
    shipment = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not shipment:
        raise HTTPException(status_code=404, detail="物流记录不存在")
    order = db.query(Order).filter(Order.id == shipment.order_id).first()
    traces = db.query(ShipmentTrace).filter(ShipmentTrace.shipment_id == shipment.id).order_by(ShipmentTrace.trace_time.desc().nullslast(), ShipmentTrace.id.desc()).all()
    return {
        "shipment_id": shipment.id,
        "order_id": shipment.order_id,
        "order_no": order.order_no if order else "",
        "supplier_code": order.supplier_code if order else "",
        "courier_company": shipment.courier_company or "",
        "courier_code": shipment.courier_code or "",
        "tracking_no": shipment.tracking_no or "",
        "provider_name": shipment.provider_name or "",
        "ship_status": shipment.ship_status or "pending",
        "sync_status": getattr(shipment, 'sync_status', 'pending') or 'pending',
        "sync_error": getattr(shipment, 'sync_error', '') or '',
        "last_trace_text": shipment.last_trace_text or "",
        "last_trace_time": shipment.last_trace_time.isoformat() if shipment.last_trace_time else "",
        "last_sync_at": getattr(shipment, 'last_sync_at', None).isoformat() if getattr(shipment, 'last_sync_at', None) else "",
        "traces": [{"trace_time": t.trace_time.isoformat() if t.trace_time else "", "trace_status": t.trace_status or "", "trace_text": t.trace_text or ""} for t in traces],
    }

@router.post("/logistics/{shipment_id}/sync")
def admin_sync_logistics(shipment_id: int, db: Session = Depends(get_db)):
    shipment = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not shipment:
        raise HTTPException(status_code=404, detail="物流记录不存在")
    result = sync_one_shipment(db, shipment)
    db.commit()
    return result



@router.get("/suppliers")
def admin_suppliers(db: Session = Depends(get_db)):
    rows = db.query(Supplier).order_by(Supplier.id.asc()).all()
    return [supplier_to_dict(r) for r in rows]

@router.post("/suppliers")
def admin_save_supplier(payload: schemas.SupplierIn, db: Session = Depends(get_db)):
    supplier_code = (payload.supplier_code or "").strip()
    supplier_name = (payload.supplier_name or "").strip()
    if not supplier_code:
        raise HTTPException(status_code=400, detail="供应链编码不能为空")
    if not supplier_name:
        raise HTTPException(status_code=400, detail="供应链名称不能为空")
    exists = db.query(Supplier).filter(Supplier.supplier_code == supplier_code)
    if payload.id:
        exists = exists.filter(Supplier.id != payload.id)
    if exists.first():
        raise HTTPException(status_code=400, detail="供应链编码已存在")
    item = db.query(Supplier).filter(Supplier.id == payload.id).first() if payload.id else Supplier()
    if payload.id and not item:
        raise HTTPException(status_code=404, detail="供应链不存在")
    shipping_bot = None
    if payload.shipping_bot_code:
        shipping_bot = db.query(BotConfig).filter(BotConfig.bot_code == payload.shipping_bot_code, BotConfig.bot_type == 'shipping').first()
        if not shipping_bot:
            raise HTTPException(status_code=400, detail="绑定的供应链机器人不存在")
    old_bot_code = item.shipping_bot_code if payload.id else ''
    if not payload.id:
        db.add(item)
    data = payload.model_dump()
    data['api_key'] = _masked_or_keep(payload.api_key, item.api_key if payload.id else '')
    data['api_secret'] = _masked_or_keep(payload.api_secret, item.api_secret if payload.id else '')
    for k, v in data.items():
        if k != 'id':
            setattr(item, k, v)
    if shipping_bot:
        shipping_bot.supplier_code = supplier_code
    if old_bot_code and old_bot_code != item.shipping_bot_code:
        old_bot = db.query(BotConfig).filter(BotConfig.bot_code == old_bot_code, BotConfig.bot_type == 'shipping').first()
        if old_bot and old_bot.supplier_code == supplier_code:
            old_bot.supplier_code = ''
    db.commit()
    db.refresh(item)
    return {"ok": True, "data": supplier_to_dict(item)}

@router.post("/suppliers/{supplier_id}/toggle")
def admin_toggle_supplier(supplier_id: int, db: Session = Depends(get_db)):
    item = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="供应链不存在")
    item.is_active = not bool(item.is_active)
    db.commit()
    db.refresh(item)
    row = product_to_dict(item, db=db)
    return {"ok": True, "is_active": item.is_active, "message": ("商品已上架" if item.is_active else "商品已下架"), "item": row, "data": row}

@router.get("/product-supplier-map")
def admin_product_supplier_map(db: Session = Depends(get_db)):
    rows = db.query(ProductSupplierMap).order_by(ProductSupplierMap.product_id.asc(), ProductSupplierMap.priority.asc(), ProductSupplierMap.id.asc()).all()
    return [product_supplier_map_to_dict(r, db) for r in rows]

@router.post("/product-supplier-map")
def admin_save_product_supplier_map(payload: schemas.ProductSupplierMapIn, db: Session = Depends(get_db)):
    product = db.query(Product).filter(Product.id == payload.product_id).first()
    if not product:
        raise HTTPException(status_code=400, detail="商品不存在")
    supplier = db.query(Supplier).filter(Supplier.id == payload.supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=400, detail="供应链不存在")
    duplicate = db.query(ProductSupplierMap).filter(ProductSupplierMap.product_id == payload.product_id, ProductSupplierMap.supplier_id == payload.supplier_id)
    if payload.id:
        duplicate = duplicate.filter(ProductSupplierMap.id != payload.id)
    if duplicate.first():
        raise HTTPException(status_code=400, detail="该商品已绑定过这个供应链")
    item = db.query(ProductSupplierMap).filter(ProductSupplierMap.id == payload.id).first() if payload.id else ProductSupplierMap()
    if payload.id and not item:
        raise HTTPException(status_code=404, detail="映射不存在")
    if not payload.id:
        db.add(item)
    for k, v in payload.model_dump().items():
        if k != 'id':
            setattr(item, k, v)
    if payload.is_default:
        db.query(ProductSupplierMap).filter(ProductSupplierMap.product_id == payload.product_id, ProductSupplierMap.id != getattr(item, 'id', 0)).update({ProductSupplierMap.is_default: False})
    db.commit()
    db.refresh(item)
    return {"ok": True, "data": product_supplier_map_to_dict(item, db)}

@router.get("/order-fulfillments")
def admin_order_fulfillments(supplier_code: str | None = None, db: Session = Depends(get_db)):
    q = db.query(OrderFulfillment)
    if supplier_code:
        supplier = db.query(Supplier).filter(Supplier.supplier_code == supplier_code).first()
        q = q.filter(OrderFulfillment.supplier_id == (supplier.id if supplier else -1))
    rows = q.order_by(OrderFulfillment.id.desc()).all()
    return [order_fulfillment_to_dict(r, db) for r in rows]

@router.post("/orders/{order_id}/assign-supplier")
def admin_assign_supplier(order_id: int, payload: schemas.AssignSupplierIn, db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    supplier = db.query(Supplier).filter(Supplier.id == payload.supplier_id, Supplier.is_active == True).first()
    if not supplier:
        raise HTTPException(status_code=400, detail="供应链不存在或已停用")
    fulfillment = upsert_order_fulfillment(db, order, supplier, status='assigned', note='后台手动指派供应链')
    db.commit()
    db.refresh(fulfillment)
    return {"ok": True, "data": order_fulfillment_to_dict(fulfillment, db)}

@router.post("/orders/{order_id}/auto-assign-supplier")
def admin_auto_assign_supplier(order_id: int, db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    items = db.query(OrderItem).filter(OrderItem.order_id == order.id).all()
    product_ids = [i.product_id for i in items if i.product_id]
    supplier, reason = resolve_order_supplier(db, product_ids)
    fulfillment = upsert_order_fulfillment(db, order, supplier, status='assigned' if supplier else 'unassigned', note=reason)
    db.commit()
    data = order_fulfillment_to_dict(fulfillment, db) if fulfillment else None
    return {"ok": True, "reason": reason, "data": data, "supplier_code": order.supplier_code or ""}

@router.post("/orders/{order_id}/push-supplier")
def admin_push_supplier(order_id: int, db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    if order.pay_status != 'paid':
        raise HTTPException(status_code=400, detail="订单未支付，不能推送供应链")
    supplier = None
    if order.supplier_code:
        supplier = db.query(Supplier).filter(Supplier.supplier_code == order.supplier_code).first()
    if not supplier:
        raise HTTPException(status_code=400, detail="订单尚未分配有效供应链")
    fulfillment = db.query(OrderFulfillment).filter(OrderFulfillment.order_id == order.id, OrderFulfillment.supplier_id == supplier.id).order_by(OrderFulfillment.id.desc()).first()
    result = push_order_to_supplier(db, order, supplier, fulfillment=fulfillment, operator_name='admin')
    fulfillment = db.query(OrderFulfillment).filter(OrderFulfillment.order_id == order.id, OrderFulfillment.supplier_id == supplier.id).order_by(OrderFulfillment.id.desc()).first()
    if fulfillment:
        fulfillment.fulfillment_status = 'pushed'
        if not fulfillment.assigned_at:
            fulfillment.assigned_at = order.created_at or datetime.utcnow()
    db.commit()
    if fulfillment:
        db.refresh(fulfillment)
    return result

@router.post("/order-fulfillments/{fulfillment_id}/pull-supplier-status")
def admin_pull_supplier_status(fulfillment_id: int, db: Session = Depends(get_db)):
    fulfillment = db.query(OrderFulfillment).filter(OrderFulfillment.id == fulfillment_id).first()
    if not fulfillment:
        raise HTTPException(status_code=404, detail="履约记录不存在")
    supplier = db.query(Supplier).filter(Supplier.id == fulfillment.supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=400, detail="关联供应链不存在")
    result = pull_supplier_status(db, fulfillment, supplier)
    db.commit()
    db.refresh(fulfillment)
    return result

@router.get("/orders/{order_id}/supplier-payload")
def admin_supplier_payload_preview(order_id: int, db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    supplier = db.query(Supplier).filter(Supplier.supplier_code == order.supplier_code).first() if order.supplier_code else None
    if not supplier:
        raise HTTPException(status_code=400, detail="订单尚未分配有效供应链")
    payload = build_supplier_payload(db, order, supplier)
    return {"ok": True, "supplier_code": supplier.supplier_code, "template_type": supplier.template_type or 'standard', "payload": payload}



@router.get("/chat-overview")
def admin_chat_overview(db: Session = Depends(get_db)):
    return get_chat_overview(db)


@router.get("/chat-sessions")
def admin_chat_sessions(
    bot_code: str = "",
    status: str = "open",
    q: str = "",
    only_unread: bool = False,
    limit: int = 30,
    page: int = 1,
    page_size: int | None = None,
    db: Session = Depends(get_db),
):
    return list_sessions(
        db,
        bot_code=bot_code.strip(),
        status=(status or "open").strip(),
        q=q.strip(),
        only_unread=bool(only_unread),
        limit=limit,
        page=page,
        page_size=page_size,
    )


@router.get("/chat-sessions/{session_id}")
def admin_chat_session_detail(session_id: int, mark_read: bool = False, db: Session = Depends(get_db)):
    try:
        return get_session_detail(db, session_id, mark_read=mark_read)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.post("/chat-sessions/{session_id}/read")
def admin_chat_session_read(session_id: int, db: Session = Depends(get_db)):
    try:
        return mark_session_read(db, session_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.post("/chat-sessions/{session_id}/close")
def admin_chat_session_close(session_id: int, db: Session = Depends(get_db)):
    try:
        return close_session(db, session_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.post("/chat-sessions/{session_id}/reopen")
def admin_chat_session_reopen(session_id: int, db: Session = Depends(get_db)):
    try:
        return reopen_session(db, session_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.post("/chat-sessions/{session_id}/reply")
def admin_chat_session_reply(session_id: int, payload: schemas.ChatReplyIn, db: Session = Depends(get_db)):
    try:
        return send_session_reply(db, session_id, payload.text, operator_name=payload.operator_name or "session_bot")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/chat-events")
def admin_chat_events(payload: dict = Body(...), db: Session = Depends(get_db)):
    try:
        return record_customer_event(db, payload)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))




@router.get("/chat-keyword-blocks")
def admin_chat_keyword_blocks(db: Session = Depends(get_db)):
    return list_keyword_blocks(db)


@router.get("/chat-keyword-blocks/effective")
def admin_chat_keyword_blocks_effective(db: Session = Depends(get_db)):
    return get_effective_keyword_blocks(db)


@router.post("/chat-keyword-blocks")
def admin_save_chat_keyword_block(payload: schemas.ChatKeywordBlockIn, db: Session = Depends(get_db)):
    try:
        return save_keyword_block(db, payload.model_dump())
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/chat-keyword-blocks/{block_id}/toggle")
def admin_toggle_chat_keyword_block(block_id: int, db: Session = Depends(get_db)):
    try:
        return toggle_keyword_block(db, block_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/chat-keyword-blocks/{block_id}")
def admin_delete_chat_keyword_block(block_id: int, db: Session = Depends(get_db)):
    try:
        return delete_keyword_block(db, block_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/data-center/overview")
def admin_data_center_overview(days: int = 30, supplier_code: str | None = None, db: Session = Depends(get_db)):
    return get_data_center_overview(db, days=days, supplier_code=(supplier_code or None))


@router.get("/data-center/supplier-board")
def admin_data_center_supplier_board(days: int = 30, supplier_code: str | None = None, sort_by: str = 'paid_gmv', db: Session = Depends(get_db)):
    return get_data_center_supplier_board(db, days=days, supplier_code=(supplier_code or None), sort_by=sort_by)


@router.get("/data-center/supplier-board.xlsx")
def admin_data_center_supplier_board_xlsx(days: int = 30, supplier_code: str | None = None, sort_by: str = 'paid_gmv', db: Session = Depends(get_db)):
    content = export_data_center_supplier_board_xlsx(db, days=days, supplier_code=(supplier_code or None), sort_by=sort_by)
    filename = f"supplier_board_{days}d.xlsx"
    return StreamingResponse(io.BytesIO(content), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/data-center/trend")
def admin_data_center_trend(days: int = 30, supplier_code: str | None = None, db: Session = Depends(get_db)):
    return get_data_center_trend(db, days=days, supplier_code=(supplier_code or None))


@router.get("/data-center/category-supplier-board")
def admin_data_center_category_supplier_board(days: int = 30, supplier_code: str | None = None, db: Session = Depends(get_db)):
    return get_data_center_category_supplier_board(db, days=days, supplier_code=(supplier_code or None))


@router.get("/data-center/product-ranking")
def admin_data_center_product_ranking(days: int = 30, supplier_code: str | None = None, limit: int = 20, db: Session = Depends(get_db)):
    return get_data_center_product_ranking(db, days=days, supplier_code=(supplier_code or None), limit=limit)


@router.get("/data-center/product-ranking.xlsx")
def admin_data_center_product_ranking_xlsx(days: int = 30, supplier_code: str | None = None, limit: int = 50, db: Session = Depends(get_db)):
    content = export_data_center_product_ranking_xlsx(db, days=days, supplier_code=(supplier_code or None), limit=limit)
    filename = f"product_ranking_{days}d.xlsx"
    return StreamingResponse(io.BytesIO(content), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/data-center/funnel")
def admin_data_center_funnel(days: int = 30, supplier_code: str | None = None, db: Session = Depends(get_db)):
    return get_data_center_funnel(db, days=days, supplier_code=(supplier_code or None))


@router.get("/data-center/alerts-trend")
def admin_data_center_alerts_trend(days: int = 30, supplier_code: str | None = None, db: Session = Depends(get_db)):
    return get_data_center_alerts_trend(db, days=days, supplier_code=(supplier_code or None))

from .admin_page import render_admin_login_page, render_admin_page


@router.get("/auth/check")
def admin_auth_check(request: Request):
    configured = admin_login_configured()
    authorized = admin_request_authorized(request)
    profile = get_current_admin_profile(request) if (authorized or not configured) else {}
    return {
        "ok": authorized or not configured,
        "configured": configured,
        "authorized": authorized or not configured,
        "username": profile.get("username", "") if (authorized or not configured) else "",
        "display_name": profile.get("display_name", "") if (authorized or not configured) else "",
        "role": profile.get("role", "") if (authorized or not configured) else "",
        "is_superadmin": bool(profile.get("is_superadmin")) if (authorized or not configured) else False,
        "mode": "multi_account" if configured else "open",
    }


@router.get("/login", response_class=HTMLResponse)
def admin_login_page(request: Request):
    if admin_request_authorized(request):
        return RedirectResponse(url="/admin/ui", status_code=302)
    return render_admin_login_page(settings.admin_username or "admin")


@router.post("/login")
def admin_login(payload: dict = Body(default={}), db: Session = Depends(get_db)):
    username = str((payload or {}).get("username") or "").strip()
    password = str((payload or {}).get("password") or "").strip()
    user = authenticate_admin_user(db, username, password)
    if user:
        touch_admin_login(db, user)
        response = JSONResponse({
            "ok": True,
            "username": user.username,
            "display_name": user.display_name or user.username,
            "role": user.role or "operator",
            "is_superadmin": (user.role or "operator") == "superadmin",
        })
        set_admin_session_cookie(response, user.username)
        return response
    if not admin_login_credentials_match(username, password):
        raise HTTPException(status_code=401, detail="账号或密码错误")
    final_username = (settings.admin_username or "admin").strip() or "admin"
    response = JSONResponse({"ok": True, "username": final_username, "display_name": final_username, "role": "superadmin", "is_superadmin": True})
    set_admin_session_cookie(response, final_username)
    return response


@router.post("/logout")
def admin_logout():
    response = JSONResponse({"ok": True})
    clear_admin_session_cookie(response)
    return response


@router.get("/users")
def admin_list_users(request: Request, db: Session = Depends(get_db)):
    _require_superadmin_request(request, db)
    return list_admin_users(db)


@router.post("/users")
def admin_save_user(request: Request, payload: dict = Body(default={}), db: Session = Depends(get_db)):
    operator = _require_superadmin_request(request, db)
    user_id = (payload or {}).get("id")
    display_name = str((payload or {}).get("display_name") or "").strip()
    role = str((payload or {}).get("role") or "operator").strip() or "operator"
    is_active = bool((payload or {}).get("is_active", True))
    if user_id:
        try:
            user = update_admin_user(
                db,
                user_id=int(user_id),
                display_name=display_name,
                role=role,
                is_active=is_active,
                operator_username=operator.username,
            )
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        return {"ok": True, "mode": "updated", "user": admin_user_to_dict(user)}
    username = str((payload or {}).get("username") or "").strip()
    password = str((payload or {}).get("password") or "").strip()
    try:
        user = create_admin_user(
            db,
            username=username,
            password=password,
            display_name=display_name,
            role=role,
            is_active=is_active,
            created_by=operator.username,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"ok": True, "mode": "created", "user": admin_user_to_dict(user)}


@router.post("/users/{user_id}/password")
def admin_reset_user_password(user_id: int, request: Request, payload: dict = Body(default={}), db: Session = Depends(get_db)):
    _require_superadmin_request(request, db)
    new_password = str((payload or {}).get("new_password") or "").strip()
    try:
        user = set_admin_user_password(db, user_id=user_id, new_password=new_password)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"ok": True, "user": admin_user_to_dict(user)}


@router.post("/users/me/password")
def admin_change_my_password(request: Request, payload: dict = Body(default={}), db: Session = Depends(get_db)):
    username = get_current_admin_username(request)
    if not username:
        raise HTTPException(status_code=401, detail="后台未授权，请先登录后台")
    current_password = str((payload or {}).get("current_password") or "").strip()
    new_password = str((payload or {}).get("new_password") or "").strip()
    user = authenticate_admin_user(db, username, current_password)
    if not user:
        raise HTTPException(status_code=400, detail="当前密码不正确")
    try:
        user = set_admin_user_password(db, user_id=user.id, new_password=new_password)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"ok": True, "user": admin_user_to_dict(user)}


@router.post("/users/{user_id}/toggle")
def admin_toggle_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    operator = _require_superadmin_request(request, db)
    user = db.query(AdminUser).filter(AdminUser.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="管理员账号不存在")
    try:
        user = update_admin_user(
            db,
            user_id=user.id,
            display_name=user.display_name or user.username,
            role=user.role or "operator",
            is_active=not bool(user.is_active),
            operator_username=operator.username,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"ok": True, "user": admin_user_to_dict(user)}


@router.delete("/users/{user_id}")
def admin_delete_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    operator = _require_superadmin_request(request, db)
    try:
        delete_admin_user(db, user_id=user_id, operator_username=operator.username)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"ok": True}


@router.get("/ui", response_class=HTMLResponse)
def admin_ui(request: Request):
    return render_admin_page(get_current_admin_username(request) or (settings.admin_username or "admin"))

def _normalize_target_bot_types(values: Any) -> list[str]:
    if isinstance(values, str):
        raw = [x.strip().lower() for x in values.replace('，', ',').split(',')]
    elif isinstance(values, (list, tuple, set)):
        raw = [str(x).strip().lower() for x in values]
    else:
        raw = []
    allow = {'buyer', 'shipping', 'session'}
    rows = [x for x in raw if x in allow]
    return rows or ['buyer']


def _target_bot_types_text(values: Any) -> str:
    mapping = {'buyer': '商城机器人', 'shipping': '供应链机器人', 'session': '聚合聊天机器人'}
    rows = _normalize_target_bot_types(values)
    return '、'.join(mapping.get(x, x) for x in rows)


ANNOUNCEMENT_ALLOWED_MEDIA_MODES = {'none', 'single_video', 'video_album'}
ANNOUNCEMENT_ALLOWED_TEXT_MODES = {'caption_first', 'album_then_text'}
ANNOUNCEMENT_ALLOWED_FALLBACK = {'text_only', 'welcome_only', 'single_video_first_item', 'none'}
ANNOUNCEMENT_MAX_MEDIA_ITEMS = 4
ANNOUNCEMENT_MAX_UPLOAD_MB = 200


def _normalize_announcement_media_items(raw: Any) -> list[dict]:
    parsed = raw
    if isinstance(raw, str):
        try:
            parsed = json.loads(raw or '[]')
        except Exception:
            parsed = []
    if not isinstance(parsed, list):
        return []
    rows: list[dict] = []
    for idx, item in enumerate(parsed, start=1):
        if not isinstance(item, dict):
            continue
        url = str(item.get('normalized_url') or item.get('url') or item.get('source_url') or '').strip()
        if not url:
            continue
        media_type = str(item.get('type') or 'video').strip().lower()
        if media_type != 'video':
            continue
        sort = item.get('sort')
        try:
            sort_value = int(sort)
        except Exception:
            sort_value = idx
        rows.append({
            'type': 'video',
            'url': url,
            'source_url': str(item.get('source_url') or url).strip(),
            'normalized_url': str(item.get('normalized_url') or url).strip(),
            'sort': sort_value,
            'enabled': bool(item.get('enabled', True)),
        })
    rows = [x for x in rows if x['enabled']]
    rows.sort(key=lambda x: (int(x.get('sort') or 9999), x.get('url') or ''))
    return rows[:ANNOUNCEMENT_MAX_MEDIA_ITEMS]


def _validate_announcement_media_items(media_mode: str, media_items: list[dict], text_mode: str, content_text: str):
    if media_mode not in ANNOUNCEMENT_ALLOWED_MEDIA_MODES:
        raise HTTPException(status_code=400, detail='公告媒体模式不正确')
    if text_mode not in ANNOUNCEMENT_ALLOWED_TEXT_MODES:
        raise HTTPException(status_code=400, detail='公告文案模式不正确')
    if media_mode == 'video_album':
        if len(media_items) < 2:
            raise HTTPException(status_code=400, detail='相册组模式至少需要 2 个视频')
        if len(media_items) > ANNOUNCEMENT_MAX_MEDIA_ITEMS:
            raise HTTPException(status_code=400, detail=f'相册组模式最多 {ANNOUNCEMENT_MAX_MEDIA_ITEMS} 个视频')
        seen_sorts = set()
        for row in media_items:
            if row.get('type') != 'video':
                raise HTTPException(status_code=400, detail='相册组仅支持视频')
            if not str(row.get('url') or '').strip():
                raise HTTPException(status_code=400, detail='相册组视频地址不能为空')
            sort_value = int(row.get('sort') or 0)
            if sort_value in seen_sorts:
                raise HTTPException(status_code=400, detail='相册组视频排序不能重复')
            seen_sorts.add(sort_value)
    if text_mode == 'caption_first' and len(str(content_text or '').strip()) > 1024:
        raise HTTPException(status_code=400, detail='短文案过长，Telegram caption 最多 1024 字')


def _announcement_mode_from_item(item: AnnouncementConfig | None) -> str:
    if not item:
        return 'none'
    mode = str(getattr(item, 'media_mode', '') or '').strip().lower()
    if mode in ANNOUNCEMENT_ALLOWED_MEDIA_MODES:
        return mode
    media_url = str(getattr(item, 'media_url', '') or '').strip()
    media_type = str(getattr(item, 'media_type', '') or '').strip().lower()
    if media_type == 'video' and media_url:
        return 'single_video'
    return 'none'


def _announcement_to_dict(item: AnnouncementConfig | None) -> dict:
    if not item:
        return {
            'scene': 'startup',
            'title': '',
            'content_text': '',
            'media_type': 'none',
            'media_url': '',
            'media_mode': 'none',
            'media_items': [],
            'media_cache': [],
            'media_normalize_status': 'pending',
            'media_normalize_error': '',
            'media_cache_updated_at': '',
            'text_mode': 'caption_first',
            'target_bot_types': ['buyer'],
            'is_enabled': False,
            'send_caption': True,
            'replace_start_welcome': True,
            'fallback_mode': 'text_only',
            'updated_at': '',
        }
    media_mode = _announcement_mode_from_item(item)
    media_items = _normalize_announcement_media_items(getattr(item, 'media_items_json', '[]'))
    media_cache = pick_album_send_items(getattr(item, 'media_cache_json', '[]'), getattr(item, 'media_items_json', '[]'))
    media_url = str(item.media_url or '').strip()
    if media_mode == 'single_video' and media_url and not media_items:
        media_items = [{'type': 'video', 'url': media_url, 'sort': 1, 'enabled': True}]
    return {
        'id': item.id,
        'scene': item.scene,
        'title': item.title or '',
        'content_text': item.content_text or '',
        'media_type': item.media_type or 'none',
        'media_url': media_url,
        'media_mode': media_mode,
        'media_items': media_items,
        'media_cache': media_cache,
        'media_normalize_status': str(getattr(item, 'media_normalize_status', '') or 'pending'),
        'media_normalize_error': str(getattr(item, 'media_normalize_error', '') or ''),
        'media_cache_updated_at': item.media_cache_updated_at.isoformat() if getattr(item, 'media_cache_updated_at', None) else '',
        'text_mode': str(getattr(item, 'text_mode', '') or 'caption_first'),
        'target_bot_types': _normalize_target_bot_types(item.target_bot_types),
        'target_bot_types_text': _target_bot_types_text(item.target_bot_types),
        'is_enabled': bool(item.is_enabled),
        'send_caption': bool(item.send_caption),
        'replace_start_welcome': bool(getattr(item, 'replace_start_welcome', True)),
        'fallback_mode': str(getattr(item, 'fallback_mode', '') or 'text_only'),
        'updated_at': item.updated_at.isoformat() if item.updated_at else '',
    }


def _clear_announcement_media_cache_json(raw: str, sorts: list[int] | None = None) -> tuple[str, list[int]]:
    try:
        rows = json.loads(raw or '[]')
        if not isinstance(rows, list):
            rows = []
    except Exception:
        rows = []
    target_sorts = {int(x) for x in (sorts or []) if str(x).strip().isdigit()}
    clear_all = not target_sorts
    cleared = []
    new_rows = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        current_sort = int(row.get('sort') or 0)
        new_row = dict(row)
        if clear_all or current_sort in target_sorts:
            new_row['telegram_file_id'] = ''
            new_row['telegram_unique_id'] = ''
            new_row['error'] = ''
            new_row['updated_at'] = datetime.utcnow().isoformat()
            cleared.append(current_sort)
        new_rows.append(new_row)
    return json.dumps(new_rows, ensure_ascii=False), sorted(x for x in cleared if x > 0)


def _send_telegram_media(bot_token: str, chat_id: str, media_type: str, media_url: str, text: str):
    text = str(text or "").strip()
    media_type = str(media_type or "none").strip().lower()
    media_url = str(media_url or "").strip()

    if media_type == "video" and media_url:
        internal_url = _internal_media_url(media_url)
        print(f"[announcement] sendVideo start chat_id={chat_id} src={media_url} internal={internal_url}")
        with httpx.Client(timeout=120.0) as client:
            res = client.get(internal_url)
            res.raise_for_status()
            filename = Path(internal_url.split("?", 1)[0]).name or "announcement_video.mp4"
            mime = res.headers.get("content-type") or "video/mp4"
            file_size = len(res.content or b"")
            print(f"[announcement] sendVideo fetched ok chat_id={chat_id} file={filename} size={file_size} mime={mime}")
            data = {"chat_id": str(chat_id)}
            if text:
                data["caption"] = text[:1024]
            result = _telegram_multipart(
                bot_token,
                "sendVideo",
                data=data,
                files={"video": (filename, res.content, mime)},
            )
            print(f"[announcement] sendVideo success chat_id={chat_id} file={filename}")
            return result

    print(f"[announcement] sendMessage fallback chat_id={chat_id} text_len={len(text)}")
    return _telegram_json(bot_token, "sendMessage", {"chat_id": str(chat_id), "text": text or "商城公告"})


def _send_telegram_media_group(bot_token: str, chat_id: str, media_items: list[dict], text: str):
    if len(media_items) < 2:
        raise RuntimeError('相册组发送至少需要 2 个视频')
    payload_items = []
    file_map: dict[str, tuple[str, bytes, str]] = {}
    with httpx.Client(timeout=180.0) as client:
        for idx, item in enumerate(media_items[:ANNOUNCEMENT_MAX_MEDIA_ITEMS]):
            src_url = str(item.get('url') or '').strip()
            internal_url = _internal_media_url(src_url)
            print(f"[announcement] sendMediaGroup fetch chat_id={chat_id} idx={idx} src={src_url} internal={internal_url}")
            res = client.get(internal_url)
            res.raise_for_status()
            filename = Path(internal_url.split('?', 1)[0]).name or f"announcement_video_{idx+1}.mp4"
            mime = res.headers.get('content-type') or 'video/mp4'
            attach_name = f"file{idx}"
            file_map[attach_name] = (filename, res.content, mime)
            row = {'type': 'video', 'media': f'attach://{attach_name}'}
            if idx == 0 and text:
                row['caption'] = text[:1024]
            payload_items.append(row)
    data = {'chat_id': str(chat_id), 'media': json.dumps(payload_items, ensure_ascii=False)}
    result = _telegram_multipart(bot_token, 'sendMediaGroup', data=data, files=file_map)
    print(f"[announcement] sendMediaGroup success chat_id={chat_id} count={len(payload_items)}")
    return result


def _find_announcement_config(db: Session, scene: str = 'startup') -> AnnouncementConfig | None:
    return db.query(AnnouncementConfig).filter(AnnouncementConfig.scene == scene).first()


def _upsert_announcement_config(db: Session, scene: str, payload: dict) -> AnnouncementConfig:
    item = _find_announcement_config(db, scene)
    if not item:
        item = AnnouncementConfig(scene=scene)
        db.add(item)

    title = str(payload.get('title') or '').strip()
    content_text = str(payload.get('content_text') or '').strip()
    media_mode = str(payload.get('media_mode') or '').strip().lower()
    if media_mode not in ANNOUNCEMENT_ALLOWED_MEDIA_MODES:
        media_mode = 'single_video' if str(payload.get('media_url') or '').strip() else 'none'
    text_mode = str(payload.get('text_mode') or 'caption_first').strip().lower()
    if text_mode not in ANNOUNCEMENT_ALLOWED_TEXT_MODES:
        text_mode = 'caption_first'
    fallback_mode = str(payload.get('fallback_mode') or 'text_only').strip().lower()
    if fallback_mode not in ANNOUNCEMENT_ALLOWED_FALLBACK:
        fallback_mode = 'text_only'
    replace_start_welcome = bool(payload.get('replace_start_welcome', True))
    media_items = _normalize_announcement_media_items(payload.get('media_items'))
    media_url = str(payload.get('media_url') or '').strip()

    if media_mode == 'single_video':
        if not media_url and media_items:
            media_url = str(media_items[0].get('url') or '').strip()
        media_items = ([{'type': 'video', 'url': media_url, 'sort': 1, 'enabled': True}] if media_url else [])
    elif media_mode == 'none':
        media_items = []
        media_url = ''
    elif media_mode == 'video_album' and not media_items and media_url:
        media_items = [{'type': 'video', 'url': media_url, 'sort': 1, 'enabled': True}]

    _validate_announcement_media_items(media_mode, media_items, text_mode, content_text)

    item.title = title
    item.content_text = content_text
    item.media_mode = media_mode
    item.media_items_json = json.dumps(media_items, ensure_ascii=False)
    item.media_cache_json = json.dumps(merge_media_cache(getattr(item, 'media_cache_json', '[]'), media_items), ensure_ascii=False)
    item.media_normalize_status = 'ready' if media_items else 'pending'
    item.media_normalize_error = ''
    item.media_cache_updated_at = datetime.utcnow() if media_items else None
    item.text_mode = text_mode
    item.target_bot_types = ','.join(_normalize_target_bot_types(payload.get('target_bot_types')))
    item.is_enabled = bool(payload.get('is_enabled'))
    item.send_caption = bool(payload.get('send_caption', True))
    item.replace_start_welcome = replace_start_welcome
    item.fallback_mode = fallback_mode
    item.media_type = 'video' if media_mode in {'single_video', 'video_album'} and media_items else 'none'
    item.media_url = media_url or (str(media_items[0].get('normalized_url') or media_items[0].get('url') or '').strip() if media_items else '')
    item.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(item)
    return item


def _get_bot_rows_for_types(db: Session, target_bot_types: list[str]) -> list[BotConfig]:
    q = db.query(BotConfig).filter(BotConfig.is_enabled == True)
    if target_bot_types:
        q = q.filter(BotConfig.bot_type.in_(target_bot_types))
    return q.order_by(BotConfig.id.asc()).all()


def _send_announcement_by_config(bot_token: str, chat_id: str, cfg: AnnouncementConfig):
    data = _announcement_to_dict(cfg)
    text = str(data.get('content_text') or '').strip()
    media_mode = str(data.get('media_mode') or 'none').strip().lower()
    media_items = _normalize_announcement_media_items(data.get('media_items'))
    if media_mode == 'video_album' and len(media_items) >= 2:
        return _send_telegram_media_group(bot_token, chat_id, media_items, text)
    if media_mode in {'single_video', 'video_album'} and media_items:
        return _send_telegram_media(bot_token, chat_id, 'video', str(media_items[0].get('url') or ''), text)
    return _send_telegram_media(bot_token, chat_id, 'none', '', text)


def _broadcast_announcement(db: Session, scene: str, target_bot_types: list[str], operator_name: str = 'admin_ui') -> dict:
    cfg = _find_announcement_config(db, scene)
    if not cfg or not cfg.is_enabled:
        raise HTTPException(status_code=400, detail='请先启用公告配置。')
    bot_rows = _get_bot_rows_for_types(db, target_bot_types)
    bot_map = {row.bot_code: row for row in bot_rows if row.bot_token}
    if not bot_map:
        raise HTTPException(status_code=400, detail='选定类型下没有可用机器人。')
    sessions = (
        db.query(CustomerSession)
        .filter(CustomerSession.session_status.in_(['open', 'closed']))
        .filter(CustomerSession.bot_code.in_(list(bot_map.keys())))
        .filter(CustomerSession.telegram_chat_id != '')
        .order_by(CustomerSession.id.asc())
        .all()
    )
    dedup = set()
    results = []
    for sess in sessions:
        key = (sess.bot_code, sess.telegram_chat_id)
        if key in dedup:
            continue
        dedup.add(key)
        bot_row = bot_map.get(sess.bot_code)
        if not bot_row or not bot_row.bot_token:
            continue
        try:
            print(f"[announcement] broadcast sending bot_code={sess.bot_code} chat_id={sess.telegram_chat_id} media_mode={_announcement_mode_from_item(cfg)}")
            _send_announcement_by_config(bot_row.bot_token, sess.telegram_chat_id, cfg)
            results.append({'bot_code': sess.bot_code, 'chat_id': sess.telegram_chat_id, 'status': 'sent'})
        except Exception as e:
            print(f"[announcement] broadcast failed bot_code={sess.bot_code} chat_id={sess.telegram_chat_id} error={e}")
            results.append({'bot_code': sess.bot_code, 'chat_id': sess.telegram_chat_id, 'status': 'failed', 'error': str(e)[:200]})
    success_count = sum(1 for x in results if x['status'] == 'sent')
    failed_count = sum(1 for x in results if x['status'] == 'failed')
    return {'scene': scene, 'target_bot_types': target_bot_types, 'total': len(results), 'success_count': success_count, 'failed_count': failed_count, 'results': results}


@router.get("/announcements/config")
def admin_get_announcement_config(scene: str = 'startup', db: Session = Depends(get_db)):
    return _announcement_to_dict(_find_announcement_config(db, scene))


@router.post("/announcements/config")
def admin_save_announcement_config(payload: dict = Body(...), db: Session = Depends(get_db)):
    scene = str((payload or {}).get('scene') or 'startup').strip() or 'startup'
    item = _upsert_announcement_config(db, scene, payload or {})
    return {'ok': True, 'data': _announcement_to_dict(item)}


@router.post("/announcements/upload-video")
async def admin_upload_announcement_video(file: UploadFile = File(...)):
    filename_src = file.filename or ''
    suffix = Path(filename_src).suffix.lower()
    if not is_supported_video_filename(filename_src):
        raise HTTPException(status_code=400, detail='公告视频格式不支持，请上传 mov/mp4/m4v/avi/mkv/webm/ts/mpeg/mpg/3gp')
    data = await file.read()
    size_mb = len(data) / 1024 / 1024
    if size_mb > ANNOUNCEMENT_MAX_UPLOAD_MB:
        raise HTTPException(status_code=400, detail=f'公告视频不能超过 {ANNOUNCEMENT_MAX_UPLOAD_MB}MB')
    raw_filename = f"announcement_video_{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}{suffix}"
    raw_dest = UPLOAD_DIR / raw_filename
    raw_dest.write_bytes(data)
    source_url = f"/static/uploads/{raw_filename}"
    normalized = normalize_announcement_video(str(raw_dest), str(UPLOAD_DIR))
    normalized_url = source_url
    normalize_status = 'ready'
    normalize_error = ''
    if normalized.get('ok') and normalized.get('normalized_url'):
        normalized_url = str(normalized.get('normalized_url') or source_url)
    else:
        normalize_status = 'failed'
        normalize_error = str(normalized.get('error') or '')[:500]
    public_url = f"{settings.backend_public_url}{normalized_url}" if normalized_url.startswith('/') else normalized_url
    return {
        'ok': True,
        'url': public_url,
        'source_url': source_url,
        'normalized_url': normalized_url,
        'normalize_status': normalize_status,
        'normalize_error': normalize_error,
        'size': len(data),
    }


@router.post('/announcements/normalize-media')
def admin_normalize_announcement_media(payload: dict = Body(...), db: Session = Depends(get_db)):
    scene = str((payload or {}).get('scene') or 'startup').strip() or 'startup'
    item = _find_announcement_config(db, scene)
    if not item:
        raise HTTPException(status_code=404, detail='公告配置不存在')
    media_items = _normalize_announcement_media_items(getattr(item, 'media_items_json', '[]'))
    if not media_items:
        raise HTTPException(status_code=400, detail='请先保存公告视频')
    cache_rows = []
    errors = []
    for row in media_items:
        src_url = str(row.get('source_url') or row.get('url') or '').strip()
        rel_path = src_url.replace('/static/uploads/', '') if src_url.startswith('/static/uploads/') else Path(src_url).name
        src_path = UPLOAD_DIR / rel_path
        normalized = normalize_announcement_video(str(src_path), str(UPLOAD_DIR))
        normalized_url = str(normalized.get('normalized_url') or src_url)
        ok = bool(normalized.get('ok'))
        if not ok:
            errors.append(str(normalized.get('error') or ''))
        cache_rows.append({
            'sort': int(row.get('sort') or 0),
            'source_url': src_url,
            'normalized_url': normalized_url,
            'telegram_file_id': '',
            'telegram_unique_id': '',
            'status': 'ready' if ok else 'failed',
            'error': '' if ok else str(normalized.get('error') or '')[:500],
            'updated_at': datetime.utcnow().isoformat(),
        })
    item.media_cache_json = json.dumps(cache_rows, ensure_ascii=False)
    item.media_normalize_status = 'failed' if errors else 'ready'
    item.media_normalize_error = '; '.join(x for x in errors if x)[:1000]
    item.media_cache_updated_at = datetime.utcnow()
    item.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(item)
    return {'ok': True, 'data': _announcement_to_dict(item)}


@router.get('/announcements/media-cache')
def admin_get_announcement_media_cache(scene: str = 'startup', db: Session = Depends(get_db)):
    item = _find_announcement_config(db, scene)
    if not item:
        return {'ok': True, 'scene': scene, 'media_cache': [], 'media_normalize_status': 'pending', 'media_normalize_error': ''}
    data = _announcement_to_dict(item)
    return {
        'ok': True,
        'scene': scene,
        'media_cache': data.get('media_cache') or [],
        'media_normalize_status': data.get('media_normalize_status') or 'pending',
        'media_normalize_error': data.get('media_normalize_error') or '',
        'media_cache_updated_at': data.get('media_cache_updated_at') or '',
    }


@router.post('/announcements/media-cache')
def admin_update_announcement_media_cache(payload: dict = Body(...), db: Session = Depends(get_db)):
    scene = str((payload or {}).get('scene') or 'startup').strip() or 'startup'
    rows = (payload or {}).get('media_file_ids') or []
    item = _find_announcement_config(db, scene)
    if not item:
        raise HTTPException(status_code=404, detail='公告配置不存在')
    item.media_cache_json = save_telegram_file_ids(getattr(item, 'media_cache_json', '[]'), rows if isinstance(rows, list) else [])
    item.media_cache_updated_at = datetime.utcnow()
    item.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(item)
    return {'ok': True, 'data': _announcement_to_dict(item)}


@router.post('/announcements/cache/clear')
def admin_clear_announcement_cache(payload: dict = Body(...), db: Session = Depends(get_db)):
    scene = str((payload or {}).get('scene') or 'startup').strip() or 'startup'
    item = _find_announcement_config(db, scene)
    if not item:
        raise HTTPException(status_code=404, detail='公告配置不存在')
    updated_json, cleared_sorts = _clear_announcement_media_cache_json(getattr(item, 'media_cache_json', '[]'), None)
    item.media_cache_json = updated_json
    item.media_cache_updated_at = datetime.utcnow()
    item.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(item)
    return {'ok': True, 'scene': scene, 'cleared_sorts': cleared_sorts, 'cleared_count': len(cleared_sorts), 'data': _announcement_to_dict(item)}


@router.post('/announcements/cache/clear-selected')
def admin_clear_selected_announcement_cache(payload: dict = Body(...), db: Session = Depends(get_db)):
    scene = str((payload or {}).get('scene') or 'startup').strip() or 'startup'
    sorts_raw = (payload or {}).get('sorts') or []
    sorts = []
    for x in sorts_raw if isinstance(sorts_raw, list) else []:
        try:
            val = int(x)
        except Exception:
            continue
        if 1 <= val <= 4:
            sorts.append(val)
    sorts = sorted(set(sorts))
    if not sorts:
        raise HTTPException(status_code=400, detail='请至少选择一个视频缓存')
    item = _find_announcement_config(db, scene)
    if not item:
        raise HTTPException(status_code=404, detail='公告配置不存在')
    updated_json, cleared_sorts = _clear_announcement_media_cache_json(getattr(item, 'media_cache_json', '[]'), sorts)
    item.media_cache_json = updated_json
    item.media_cache_updated_at = datetime.utcnow()
    item.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(item)
    return {'ok': True, 'scene': scene, 'cleared_sorts': cleared_sorts, 'cleared_count': len(cleared_sorts), 'data': _announcement_to_dict(item)}


@router.post("/announcements/broadcast")
def admin_broadcast_announcement(payload: dict = Body(...), db: Session = Depends(get_db), username: str = Depends(get_current_admin_username)):
    scene = str((payload or {}).get('scene') or 'startup').strip() or 'startup'
    target_bot_types = _normalize_target_bot_types((payload or {}).get('target_bot_types'))
    return _broadcast_announcement(db, scene, target_bot_types, operator_name=username or 'admin_ui')


@router.get("/announcements/startup")
def admin_get_startup_announcement(bot_code: str, telegram_chat_id: str, telegram_user_id: str = '', db: Session = Depends(get_db)):
    bot = db.query(BotConfig).filter(BotConfig.bot_code == bot_code).first()
    if not bot:
        print(f"[announcement] startup skip bot_missing bot_code={bot_code}")
        return {'should_send': False}
    cfg = _find_announcement_config(db, 'startup')
    if not cfg or not cfg.is_enabled:
        print(f"[announcement] startup skip cfg_disabled bot_code={bot_code}")
        return {'should_send': False}
    target_types = _normalize_target_bot_types(cfg.target_bot_types)
    if bot.bot_type not in target_types:
        print(f"[announcement] startup skip bot_type_filtered bot_code={bot_code} bot_type={bot.bot_type} target={target_types}")
        return {'should_send': False}
    receipt = db.query(AnnouncementReceipt).filter(AnnouncementReceipt.scene=='startup', AnnouncementReceipt.bot_code==bot_code, AnnouncementReceipt.telegram_chat_id==str(telegram_chat_id)).first()
    if receipt:
        print(f"[announcement] startup skip already_sent bot_code={bot_code} chat_id={telegram_chat_id}")
        return {'should_send': False}
    payload = _announcement_to_dict(cfg)
    payload.update({'should_send': True, 'scene': 'startup'})
    print(f"[announcement] startup allow bot_code={bot_code} chat_id={telegram_chat_id} media_mode={payload.get('media_mode')} items={len(payload.get('media_items') or [])}")
    return payload


@router.post('/announcements/receipt')
def admin_confirm_announcement_receipt(payload: dict = Body(...), db: Session = Depends(get_db)):
    scene = str((payload or {}).get('scene') or 'startup').strip() or 'startup'
    bot_code = str((payload or {}).get('bot_code') or '').strip()
    telegram_chat_id = str((payload or {}).get('telegram_chat_id') or '').strip()
    telegram_user_id = str((payload or {}).get('telegram_user_id') or '').strip()
    target_bot_type = str((payload or {}).get('target_bot_type') or 'buyer').strip() or 'buyer'
    if not bot_code or not telegram_chat_id:
        raise HTTPException(status_code=400, detail='bot_code 和 telegram_chat_id 不能为空')
    receipt = db.query(AnnouncementReceipt).filter(
        AnnouncementReceipt.scene == scene,
        AnnouncementReceipt.bot_code == bot_code,
        AnnouncementReceipt.telegram_chat_id == telegram_chat_id,
    ).first()
    if receipt:
        return {'ok': True, 'already_exists': True}
    receipt = AnnouncementReceipt(
        scene=scene,
        bot_code=bot_code,
        telegram_chat_id=telegram_chat_id,
        telegram_user_id=telegram_user_id,
        target_bot_type=target_bot_type,
    )
    db.add(receipt)
    db.commit()
    return {'ok': True, 'already_exists': False}
