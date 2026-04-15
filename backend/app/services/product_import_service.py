
from __future__ import annotations

import io
import json
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..models import Product, ProductCategory, ProductImportBatch, ProductImportError, ProductSupplierMap, Supplier

HEADERS = [
    "商品ID", "分类名称", "商品名称", "副标题", "SKU", "封面图地址", "售价", "原价", "库存", "重量(g)",
    "单位", "描述", "详情", "状态", "排序", "供应链编码", "供应链SKU", "供应链优先级", "默认供应链",
]

HEADER_ALIASES = {
    "商品id": "商品ID",
    "商品Id": "商品ID",
    "分类": "分类名称",
    "商品名称": "商品名称",
    "名称": "商品名称",
    "副标题": "副标题",
    "sku": "SKU",
    "sku_code": "SKU",
    "封面图": "封面图地址",
    "封面图url": "封面图地址",
    "图片地址": "封面图地址",
    "售价": "售价",
    "价格": "售价",
    "原价": "原价",
    "库存": "库存",
    "重量": "重量(g)",
    "单位": "单位",
    "描述": "描述",
    "详情": "详情",
    "detail_html": "详情",
    "状态": "状态",
    "是否启用": "状态",
    "排序": "排序",
    "供应链": "供应链编码",
    "供应链编码": "供应链编码",
    "供应链sku": "供应链SKU",
    "供应链优先级": "供应链优先级",
    "默认供应链": "默认供应链",
}

SAMPLE_ROW = {
    "商品ID": "",
    "分类名称": "新品专区",
    "商品名称": "轻薄休闲外套",
    "副标题": "春季新款 / 两色可选",
    "SKU": "JKT-001",
    "封面图地址": "https://example.com/jkt-001.jpg",
    "售价": "199",
    "原价": "299",
    "库存": "88",
    "重量(g)": "420",
    "单位": "件",
    "描述": "基础描述，支持纯文本。",
    "详情": "<p>这里可以放详情 HTML</p>",
    "状态": "启用",
    "排序": "100",
    "供应链编码": "A",
    "供应链SKU": "A-JKT-001",
    "供应链优先级": "100",
    "默认供应链": "是",
}


def _text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _to_int(v: Any, default: int = 0) -> int:
    s = _text(v)
    if not s:
        return default
    return int(Decimal(s))


def _to_decimal(v: Any, default: str = "0") -> Decimal:
    s = _text(v)
    if not s:
        return Decimal(default)
    try:
        return Decimal(s)
    except InvalidOperation as e:
        raise ValueError(f"数字格式不正确：{s}") from e


def _to_bool(v: Any, default: bool = True) -> bool:
    s = _text(v).lower()
    if not s:
        return default
    return s in {"1", "true", "yes", "y", "启用", "是", "on"}


def _normalize_status(v: Any) -> bool:
    s = _text(v)
    if not s:
        return True
    if s in {"启用", "是", "true", "1", "on", "yes"}:
        return True
    if s in {"停用", "否", "false", "0", "off", "no"}:
        return False
    raise ValueError("状态仅支持：启用/停用/是/否")


def _image_check(url: str) -> str | None:
    url = (url or "").strip()
    if not url:
        return None
    if not (url.startswith("http://") or url.startswith("https://")):
        return "图片地址格式异常，建议填写 http/https 地址"
    try:
        with httpx.Client(follow_redirects=True, timeout=5.0) as client:
            res = client.head(url)
            if res.status_code >= 400 or (res.status_code == 405):
                res = client.get(url)
            if res.status_code >= 400:
                return f"图片链接不可达：HTTP {res.status_code}"
    except Exception as e:
        return f"图片链接检测失败：{e}"
    return None


def build_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "商品导入模板"
    note_fill = PatternFill("solid", fgColor="1D4ED8")
    header_fill = PatternFill("solid", fgColor="E0E7FF")
    thin = Side(style="thin", color="D1D5DB")
    ws.merge_cells("A1:S1")
    ws["A1"] = "商品模板导入 v3：支持预览校验、正式导入、失败行二次导入、图片链接可达性检测"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=12)
    ws["A1"].fill = note_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    for idx, head in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=idx, value=head)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for idx, head in enumerate(HEADERS, start=1):
        ws.cell(row=3, column=idx, value=SAMPLE_ROW.get(head, ""))
    widths = {
        "A": 10, "B": 16, "C": 22, "D": 22, "E": 18, "F": 38, "G": 10, "H": 10, "I": 10,
        "J": 12, "K": 10, "L": 26, "M": 28, "N": 10, "O": 10, "P": 14, "Q": 18, "R": 14, "S": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A3"
    guide = wb.create_sheet("填写说明")
    guide["A1"] = "填写说明"
    guide["A1"].font = Font(bold=True, size=14)
    tips = [
        "1. 商品ID 留空表示新增；填写商品ID可更新已有商品。",
        "2. 若商品ID 留空，但 SKU 命中已有商品，也会按 SKU 更新。",
        "3. 更新时空白字段默认保留旧值。",
        "4. 分类不存在时会自动创建。",
        "5. 供应链编码存在时，会自动维护商品供应链映射。",
        "6. 默认供应链支持：是/否、true/false、1/0。",
        "7. 可先点“预览校验”，确认无误后再正式导入。",
    ]
    for i, tip in enumerate(tips, start=3):
        guide[f"A{i}"] = tip
    guide.column_dimensions["A"].width = 80
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _parse_rows(file_bytes: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    header_row = [_text(c.value) for c in ws[2]] if ws.max_row >= 2 else []
    if not any(header_row):
        header_row = [_text(c.value) for c in ws[1]]
        data_start = 2
    else:
        data_start = 3
    canon_headers = []
    for h in header_row:
        key = HEADER_ALIASES.get(h, h)
        canon_headers.append(key)
    missing = [h for h in ["商品名称", "售价"] if h not in canon_headers and "商品ID" not in canon_headers and "SKU" not in canon_headers]
    if len(canon_headers) < 5:
        raise ValueError("模板表头不正确，请下载最新模板后填写")
    rows: list[dict[str, Any]] = []
    for row_no in range(data_start, ws.max_row + 1):
        values = [ws.cell(row=row_no, column=i + 1).value for i in range(len(canon_headers))]
        if not any(v not in (None, "") for v in values):
            continue
        row = {canon_headers[i]: values[i] if i < len(values) else None for i in range(len(canon_headers))}
        row["__row_no__"] = row_no
        rows.append(row)
    return rows


def _evaluate_rows(db: Session, rows: list[dict[str, Any]], check_images: bool = False) -> dict[str, Any]:
    suppliers = {s.supplier_code.strip(): s for s in db.query(Supplier).all()}
    categories = {c.name.strip(): c for c in db.query(ProductCategory).all() if (c.name or '').strip()}
    seen_skus: set[str] = set()
    result_rows = []
    summary = {"total": 0, "ok": 0, "warn": 0, "error": 0, "create": 0, "update": 0}
    for raw in rows:
        row_no = raw.get("__row_no__", 0)
        errors: list[str] = []
        warnings: list[str] = []
        product_id = _text(raw.get("商品ID"))
        sku_code = _text(raw.get("SKU"))
        product_name = _text(raw.get("商品名称"))
        category_name = _text(raw.get("分类名称"))
        supplier_code = _text(raw.get("供应链编码"))
        cover_image = _text(raw.get("封面图地址"))
        if not product_id and not sku_code and not product_name:
            errors.append("商品ID、SKU、商品名称至少填写一项")
        product = None
        if product_id:
            if not product_id.isdigit():
                errors.append("商品ID 必须是数字")
            else:
                product = db.query(Product).filter(Product.id == int(product_id)).first()
                if not product:
                    errors.append("商品ID 不存在")
        if sku_code:
            if sku_code in seen_skus:
                errors.append("模板内 SKU 重复")
            else:
                seen_skus.add(sku_code)
            product_by_sku = db.query(Product).filter(Product.sku_code == sku_code).first()
            if product is None:
                product = product_by_sku
            elif product_by_sku and product_by_sku.id != product.id:
                errors.append("商品ID 与 SKU 对应的商品不一致")
        action = "update" if product else "create"
        if action == "create" and not product_name:
            errors.append("新增商品必须填写商品名称")
        try:
            price_cny = _to_decimal(raw.get("售价"), default="0")
            if action == "create" and price_cny <= 0:
                errors.append("新增商品时，售价必须大于 0")
        except Exception as e:
            errors.append(str(e))
        try:
            if _text(raw.get("原价")):
                _to_decimal(raw.get("原价"), default="0")
        except Exception as e:
            errors.append(f"原价{e}")
        for field in ["库存", "重量(g)", "排序", "供应链优先级"]:
            try:
                if _text(raw.get(field)):
                    _to_int(raw.get(field), 0)
            except Exception:
                errors.append(f"{field} 必须是数字")
        if category_name and category_name not in categories:
            warnings.append("分类不存在，导入时会自动创建")
        if supplier_code:
            if supplier_code not in suppliers:
                errors.append("供应链编码不存在")
        if check_images:
            msg = _image_check(cover_image)
            if msg:
                warnings.append(msg)
        elif cover_image and not (cover_image.startswith('http://') or cover_image.startswith('https://') or cover_image.startswith('/static/')):
            warnings.append("图片地址建议使用 http/https 或 /static/uploads/ 路径")
        if action == "update":
            warnings.append("更新行留空字段将保留旧值")
        status = "error" if errors else ("warn" if warnings else "ok")
        summary["total"] += 1
        summary[status] += 1
        summary[action] += 1
        result_rows.append({
            "row_no": row_no,
            "action": action,
            "status": status,
            "product_id": product.id if product else None,
            "sku_code": sku_code,
            "product_name": product_name or (product.name if product else ""),
            "category_name": category_name,
            "supplier_code": supplier_code,
            "errors": errors,
            "warnings": warnings,
            "raw": {k: (v.isoformat() if hasattr(v, 'isoformat') else v) for k, v in raw.items() if not k.startswith('__')},
        })
    return {"summary": summary, "rows": result_rows}


def preview_product_import(db: Session, file_bytes: bytes, check_images: bool = False) -> dict[str, Any]:
    rows = _parse_rows(file_bytes)
    return _evaluate_rows(db, rows, check_images=check_images)


def _upsert_category(db: Session, category_name: str) -> ProductCategory | None:
    category_name = (category_name or '').strip()
    if not category_name:
        return None
    cat = db.query(ProductCategory).filter(ProductCategory.name == category_name).first()
    if cat:
        return cat
    cat = ProductCategory(name=category_name, is_active=True, sort_order=100)
    db.add(cat)
    db.flush()
    return cat


def _sync_supplier_map(db: Session, product: Product, supplier_code: str, supplier_sku: str, priority: int, is_default: bool):
    supplier_code = (supplier_code or '').strip()
    if not supplier_code:
        return
    supplier = db.query(Supplier).filter(Supplier.supplier_code == supplier_code).first()
    if not supplier:
        raise ValueError("供应链编码不存在")
    mapping = db.query(ProductSupplierMap).filter(ProductSupplierMap.product_id == product.id, ProductSupplierMap.supplier_id == supplier.id).first()
    if not mapping:
        mapping = ProductSupplierMap(product_id=product.id, supplier_id=supplier.id)
        db.add(mapping)
    if is_default:
        db.query(ProductSupplierMap).filter(ProductSupplierMap.product_id == product.id).update({ProductSupplierMap.is_default: False})
    mapping.supplier_sku = supplier_sku
    mapping.priority = priority
    mapping.is_default = is_default
    mapping.is_active = True


def _apply_row(db: Session, raw: dict[str, Any]) -> tuple[Product, str]:
    product_id = _text(raw.get("商品ID"))
    sku_code = _text(raw.get("SKU"))
    product = None
    action = 'create'
    if product_id:
        product = db.query(Product).filter(Product.id == int(product_id)).first()
    if product is None and sku_code:
        product = db.query(Product).filter(Product.sku_code == sku_code).first()
    if product is not None:
        action = 'update'
    else:
        product = Product()
        db.add(product)
    category = _upsert_category(db, _text(raw.get("分类名称")))
    updates = {
        'category_id': category.id if category else (product.category_id if action == 'update' else None),
        'name': _text(raw.get("商品名称")),
        'subtitle': _text(raw.get("副标题")),
        'sku_code': sku_code,
        'cover_image': _text(raw.get("封面图地址")),
        'price_cny': _to_decimal(raw.get("售价"), '0'),
        'original_price_cny': _to_decimal(raw.get("原价"), '0'),
        'stock_qty': _to_int(raw.get("库存"), 0),
        'weight_gram': _to_int(raw.get("重量(g)"), 0),
        'unit_text': _text(raw.get("单位")) or '件',
        'description': _text(raw.get("描述")),
        'detail_html': _text(raw.get("详情")),
        'is_active': _normalize_status(raw.get("状态")),
        'sort_order': _to_int(raw.get("排序"), 100),
    }
    for key, value in updates.items():
        incoming_text = _text(raw.get({
            'category_id': '分类名称', 'name': '商品名称', 'subtitle': '副标题', 'sku_code': 'SKU', 'cover_image': '封面图地址',
            'price_cny': '售价', 'original_price_cny': '原价', 'stock_qty': '库存', 'weight_gram': '重量(g)', 'unit_text': '单位',
            'description': '描述', 'detail_html': '详情', 'is_active': '状态', 'sort_order': '排序'
        }.get(key, '')))
        if action == 'update' and incoming_text == '' and key not in {'is_active', 'sort_order', 'unit_text', 'stock_qty', 'weight_gram', 'price_cny', 'original_price_cny'}:
            continue
        if action == 'update' and key in {'price_cny', 'original_price_cny', 'stock_qty', 'weight_gram', 'sort_order'} and _text(raw.get({
            'price_cny':'售价','original_price_cny':'原价','stock_qty':'库存','weight_gram':'重量(g)','sort_order':'排序'}[key])) == '':
            continue
        setattr(product, key, value)
    db.flush()
    supplier_code = _text(raw.get("供应链编码"))
    if supplier_code:
        _sync_supplier_map(
            db,
            product,
            supplier_code,
            _text(raw.get("供应链SKU")),
            _to_int(raw.get("供应链优先级"), 100),
            _to_bool(raw.get("默认供应链"), False),
        )
    return product, action


def _batch_no() -> str:
    return 'P' + datetime.utcnow().strftime('%Y%m%d%H%M%S%f')


def import_product_file(db: Session, file_bytes: bytes, file_name: str, operator_name: str = '', check_images: bool = False) -> dict[str, Any]:
    parsed = preview_product_import(db, file_bytes, check_images=check_images)
    batch = ProductImportBatch(
        batch_no=_batch_no(),
        file_name=file_name or 'product_import.xlsx',
        operator_name=(operator_name or '').strip(),
        total_rows=parsed['summary']['total'],
        warning_rows=parsed['summary']['warn'],
        check_images=bool(check_images),
    )
    db.add(batch)
    db.flush()
    created_rows = updated_rows = success_rows = failed_rows = 0
    for row in parsed['rows']:
        if row['status'] == 'error':
            failed_rows += 1
            db.add(ProductImportError(
                batch_id=batch.id,
                row_no=row['row_no'],
                product_id=_text(row['raw'].get('商品ID')),
                sku_code=_text(row['raw'].get('SKU')),
                product_name=_text(row['raw'].get('商品名称')),
                error_message='；'.join(row['errors']),
                raw_row_json=json.dumps(row['raw'], ensure_ascii=False),
            ))
            continue
        try:
            product, action = _apply_row(db, row['raw'])
            success_rows += 1
            if action == 'create':
                created_rows += 1
            else:
                updated_rows += 1
        except Exception as e:
            failed_rows += 1
            db.add(ProductImportError(
                batch_id=batch.id,
                row_no=row['row_no'],
                product_id=_text(row['raw'].get('商品ID')),
                sku_code=_text(row['raw'].get('SKU')),
                product_name=_text(row['raw'].get('商品名称')),
                error_message=str(e),
                raw_row_json=json.dumps(row['raw'], ensure_ascii=False),
            ))
    batch.success_rows = success_rows
    batch.failed_rows = failed_rows
    batch.created_rows = created_rows
    batch.updated_rows = updated_rows
    db.commit()
    return {
        'ok': True,
        'batch_id': batch.id,
        'batch_no': batch.batch_no,
        'total_rows': batch.total_rows,
        'success_rows': batch.success_rows,
        'failed_rows': batch.failed_rows,
        'warning_rows': batch.warning_rows,
        'created_rows': batch.created_rows,
        'updated_rows': batch.updated_rows,
    }


def list_product_import_batches(db: Session, keyword: str = '', result_status: str = 'all') -> list[dict[str, Any]]:
    q = db.query(ProductImportBatch)
    if keyword:
        s = keyword.strip()
        q = q.filter((ProductImportBatch.batch_no.contains(s)) | (ProductImportBatch.file_name.contains(s)) | (ProductImportBatch.operator_name.contains(s)))
    rows = q.order_by(ProductImportBatch.id.desc()).all()
    data = []
    for row in rows:
        item = {
            'id': row.id,
            'batch_no': row.batch_no,
            'file_name': row.file_name,
            'operator_name': row.operator_name or '',
            'total_rows': row.total_rows,
            'success_rows': row.success_rows,
            'failed_rows': row.failed_rows,
            'warning_rows': row.warning_rows,
            'created_rows': row.created_rows,
            'updated_rows': row.updated_rows,
            'check_images': bool(row.check_images),
            'created_at': row.created_at.isoformat() if row.created_at else '',
            'result_status': 'failed' if row.failed_rows > 0 else ('warn' if row.warning_rows > 0 else 'success'),
        }
        if result_status == 'failed' and item['failed_rows'] <= 0:
            continue
        if result_status == 'warn' and item['warning_rows'] <= 0:
            continue
        if result_status == 'success' and item['failed_rows'] > 0:
            continue
        data.append(item)
    return data


def list_product_import_errors(db: Session, batch_id: int) -> list[dict[str, Any]]:
    rows = db.query(ProductImportError).filter(ProductImportError.batch_id == batch_id).order_by(ProductImportError.id.asc()).all()
    return [{
        'id': row.id,
        'batch_id': row.batch_id,
        'row_no': row.row_no,
        'product_id': row.product_id or '',
        'sku_code': row.sku_code or '',
        'product_name': row.product_name or '',
        'error_message': row.error_message or '',
        'raw_row_json': row.raw_row_json or '{}',
        'created_at': row.created_at.isoformat() if row.created_at else '',
    } for row in rows]


def build_error_workbook(errors: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = '失败行明细'
    heads = ['行号', '商品ID', 'SKU', '商品名称', '错误原因', '原始数据JSON']
    for i, h in enumerate(heads, start=1):
        ws.cell(row=1, column=i, value=h).font = Font(bold=True)
    for ridx, row in enumerate(errors, start=2):
        ws.cell(row=ridx, column=1, value=row['row_no'])
        ws.cell(row=ridx, column=2, value=row['product_id'])
        ws.cell(row=ridx, column=3, value=row['sku_code'])
        ws.cell(row=ridx, column=4, value=row['product_name'])
        ws.cell(row=ridx, column=5, value=row['error_message'])
        ws.cell(row=ridx, column=6, value=row['raw_row_json'])
    for idx, width in enumerate([10, 10, 18, 20, 48, 70], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    bio = io.BytesIO(); wb.save(bio); return bio.getvalue()


def delete_product_import_batch(db: Session, batch_id: int) -> dict[str, Any]:
    batch = db.query(ProductImportBatch).filter(ProductImportBatch.id == batch_id).first()
    if not batch:
        raise ValueError('批次不存在')
    db.query(ProductImportError).filter(ProductImportError.batch_id == batch_id).delete()
    db.delete(batch)
    db.commit()
    return {'ok': True}


def retry_product_import_batch(db: Session, batch_id: int) -> dict[str, Any]:
    batch = db.query(ProductImportBatch).filter(ProductImportBatch.id == batch_id).first()
    if not batch:
        raise ValueError('批次不存在')
    errors = db.query(ProductImportError).filter(ProductImportError.batch_id == batch_id).order_by(ProductImportError.id.asc()).all()
    if not errors:
        raise ValueError('当前批次没有失败行可重试')
    rows = []
    for err in errors:
        raw = json.loads(err.raw_row_json or '{}')
        raw['__row_no__'] = err.row_no
        rows.append(raw)
    evaluated = _evaluate_rows(db, rows, check_images=False)
    new_batch = ProductImportBatch(
        batch_no=_batch_no(),
        file_name=f'retry_{batch.file_name}',
        operator_name=batch.operator_name,
        total_rows=evaluated['summary']['total'],
        warning_rows=evaluated['summary']['warn'],
        check_images=False,
    )
    db.add(new_batch)
    db.flush()
    success_rows = failed_rows = created_rows = updated_rows = 0
    for row in evaluated['rows']:
        if row['status'] == 'error':
            failed_rows += 1
            db.add(ProductImportError(
                batch_id=new_batch.id,
                row_no=row['row_no'],
                product_id=_text(row['raw'].get('商品ID')),
                sku_code=_text(row['raw'].get('SKU')),
                product_name=_text(row['raw'].get('商品名称')),
                error_message='；'.join(row['errors']),
                raw_row_json=json.dumps(row['raw'], ensure_ascii=False),
            ))
            continue
        try:
            product, action = _apply_row(db, row['raw'])
            success_rows += 1
            created_rows += 1 if action == 'create' else 0
            updated_rows += 1 if action == 'update' else 0
        except Exception as e:
            failed_rows += 1
            db.add(ProductImportError(
                batch_id=new_batch.id,
                row_no=row['row_no'],
                product_id=_text(row['raw'].get('商品ID')),
                sku_code=_text(row['raw'].get('SKU')),
                product_name=_text(row['raw'].get('商品名称')),
                error_message=str(e),
                raw_row_json=json.dumps(row['raw'], ensure_ascii=False),
            ))
    new_batch.success_rows = success_rows
    new_batch.failed_rows = failed_rows
    new_batch.created_rows = created_rows
    new_batch.updated_rows = updated_rows
    db.commit()
    return {'ok': True, 'batch_id': new_batch.id, 'batch_no': new_batch.batch_no, 'success_rows': success_rows, 'failed_rows': failed_rows}
