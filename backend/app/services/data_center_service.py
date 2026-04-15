from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from ..models import Order, OrderItem, Product, ProductCategory, ProductSupplierMap, Shipment, Supplier


def _as_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(value)
    except Exception:
        return 0.0


def _now() -> datetime:
    return datetime.utcnow()


def _day_start(dt: datetime) -> datetime:
    return datetime(dt.year, dt.month, dt.day)


def _period(days: int) -> tuple[datetime, datetime, datetime, datetime]:
    days = max(1, min(int(days or 30), 180))
    now = _now()
    current_start = _day_start(now) - timedelta(days=days - 1)
    current_end = now
    previous_end = current_start - timedelta(microseconds=1)
    previous_start = current_start - timedelta(days=days)
    return current_start, current_end, previous_start, previous_end


def _date_range(start: datetime, end: datetime) -> list[datetime]:
    cursor = _day_start(start)
    limit = _day_start(end)
    days: list[datetime] = []
    while cursor <= limit:
        days.append(cursor)
        cursor += timedelta(days=1)
    return days


def _filter_orders(db: Session, start: datetime, end: datetime, supplier_code: str | None = None):
    q = db.query(Order).filter(Order.created_at >= start, Order.created_at <= end)
    if supplier_code:
        q = q.filter(Order.supplier_code == supplier_code)
    return q


def _shipment_rows(db: Session, order_ids: list[int]) -> list[Shipment]:
    if not order_ids:
        return []
    return db.query(Shipment).filter(Shipment.order_id.in_(order_ids)).all()


def _is_sync_abnormal(shipment: Shipment) -> bool:
    status = (shipment.sync_status or '').strip().lower()
    err = (shipment.sync_error or '').strip()
    return bool(err) or status in {'failed', 'error', 'timeout'}


def _is_stagnant(shipment: Shipment, now: datetime | None = None) -> bool:
    if (shipment.ship_status or '').strip().lower() in {'signed', 'delivered'}:
        return False
    base = shipment.last_trace_time or shipment.last_sync_at or shipment.updated_at or shipment.created_at
    now = now or _now()
    return bool(base) and (now - base) >= timedelta(hours=48)


def _is_exceptional(shipment: Shipment) -> bool:
    text = ' '.join([
        (shipment.ship_status or ''),
        (shipment.last_trace_text or ''),
        (shipment.sync_error or ''),
    ]).lower()
    return any(key in text for key in ['exception', 'returned', 'return', 'reject', 'problem', 'failed'])


def _mapped_products_count(db: Session, supplier_code: str | None = None) -> int:
    q = db.query(ProductSupplierMap).filter(ProductSupplierMap.is_active == True)
    if supplier_code:
        supplier = db.query(Supplier).filter(Supplier.supplier_code == supplier_code).first()
        if not supplier:
            return 0
        q = q.filter(ProductSupplierMap.supplier_id == supplier.id)
    return len({row.product_id for row in q.all() if row.product_id})


def _supplier_name_map(db: Session) -> dict[str, str]:
    return {s.supplier_code: s.supplier_name for s in db.query(Supplier).all() if s.supplier_code}


def _summary_from_orders(db: Session, orders: list[Order], supplier_code: str | None = None) -> dict[str, Any]:
    order_ids = [o.id for o in orders]
    shipments = _shipment_rows(db, order_ids)
    now = _now()
    paid_orders = [o for o in orders if (o.pay_status or '').lower() == 'paid']
    shipped_orders = [o for o in orders if (o.delivery_status or '').lower() in {'shipped', 'signed'}]
    signed_orders = [o for o in orders if (o.delivery_status or '').lower() == 'signed']
    pending_shipment_orders = [
        o for o in orders
        if (o.pay_status or '').lower() == 'paid' and (o.delivery_status or '').lower() not in {'shipped', 'signed'}
    ]
    pay_gmv = round(sum(_as_float(o.payable_amount) for o in paid_orders), 2)
    ship_hours = []
    for o in orders:
        if o.paid_at and o.shipped_at and o.shipped_at >= o.paid_at:
            ship_hours.append((o.shipped_at - o.paid_at).total_seconds() / 3600.0)
    avg_ship_hours = round(sum(ship_hours) / len(ship_hours), 2) if ship_hours else 0.0
    sync_abnormal_count = sum(1 for s in shipments if _is_sync_abnormal(s))
    stagnant_count = sum(1 for s in shipments if _is_stagnant(s, now=now))
    exceptional_count = sum(1 for s in shipments if _is_exceptional(s))
    sign_rate = round((len(signed_orders) / len(shipped_orders) * 100.0), 1) if shipped_orders else 0.0
    ship_rate = round((len(shipped_orders) / len(paid_orders) * 100.0), 1) if paid_orders else 0.0
    return {
        'created_orders': len(orders),
        'paid_orders': len(paid_orders),
        'paid_gmv': pay_gmv,
        'shipped_orders': len(shipped_orders),
        'signed_orders': len(signed_orders),
        'pending_shipments': len(pending_shipment_orders),
        'avg_ship_hours': avg_ship_hours,
        'sync_abnormal_count': sync_abnormal_count,
        'stagnant_count': stagnant_count,
        'exceptional_count': exceptional_count,
        'sign_rate': sign_rate,
        'ship_rate': ship_rate,
        'mapped_products_count': _mapped_products_count(db, supplier_code=supplier_code),
    }


def _with_compare(current: dict[str, Any], previous: dict[str, Any]) -> dict[str, Any]:
    metrics = {}
    for key, value in current.items():
        prev = previous.get(key, 0)
        delta = round(_as_float(value) - _as_float(prev), 2)
        pct = None
        if _as_float(prev) != 0:
            pct = round(delta / _as_float(prev) * 100.0, 1)
        metrics[key] = {
            'current': value,
            'previous': prev,
            'delta': delta,
            'delta_pct': pct,
        }
    return metrics


def get_overview(db: Session, days: int = 30, supplier_code: str | None = None) -> dict[str, Any]:
    start, end, prev_start, prev_end = _period(days)
    current_orders = _filter_orders(db, start, end, supplier_code).all()
    previous_orders = _filter_orders(db, prev_start, prev_end, supplier_code).all()
    current = _summary_from_orders(db, current_orders, supplier_code=supplier_code)
    previous = _summary_from_orders(db, previous_orders, supplier_code=supplier_code)
    metrics = _with_compare(current, previous)
    return {
        'days': days,
        'supplier_code': supplier_code or '',
        'current_start': start.isoformat(),
        'current_end': end.isoformat(),
        'previous_start': prev_start.isoformat(),
        'previous_end': prev_end.isoformat(),
        'metrics': metrics,
        'summary': {
            'current': current,
            'previous': previous,
        },
    }


def get_supplier_board(db: Session, days: int = 30, supplier_code: str | None = None, sort_by: str = 'paid_gmv') -> dict[str, Any]:
    start, end, _, _ = _period(days)
    orders = _filter_orders(db, start, end, supplier_code).all()
    supplier_names = _supplier_name_map(db)
    mapped_counts: dict[str, int] = defaultdict(int)
    for m in db.query(ProductSupplierMap).filter(ProductSupplierMap.is_active == True).all():
        supplier = db.query(Supplier).filter(Supplier.id == m.supplier_id).first()
        if not supplier or not supplier.supplier_code:
            continue
        mapped_counts[supplier.supplier_code] += 1
    buckets: dict[str, dict[str, Any]] = {}
    for s in db.query(Supplier).order_by(Supplier.supplier_code.asc()).all():
        buckets[s.supplier_code] = {
            'supplier_code': s.supplier_code,
            'supplier_name': s.supplier_name,
            'mapped_products': mapped_counts.get(s.supplier_code, 0),
            'orders': 0,
            'paid_orders': 0,
            'paid_gmv': 0.0,
            'pending_shipments': 0,
            'shipped_orders': 0,
            'signed_orders': 0,
            'ship_rate': 0.0,
            'sign_rate': 0.0,
            'avg_ship_hours': 0.0,
            'sync_abnormal_count': 0,
            'stagnant_count': 0,
            'latest_order_at': '',
        }
    shipment_cache = _shipment_rows(db, [o.id for o in orders])
    shipments_by_order: dict[int, list[Shipment]] = defaultdict(list)
    for s in shipment_cache:
        shipments_by_order[s.order_id].append(s)
    ship_hours_by_supplier: dict[str, list[float]] = defaultdict(list)
    for o in orders:
        code = (o.supplier_code or '未分配').strip() or '未分配'
        bucket = buckets.setdefault(code, {
            'supplier_code': code,
            'supplier_name': supplier_names.get(code, code if code != '未分配' else '未分配'),
            'mapped_products': mapped_counts.get(code, 0),
            'orders': 0,
            'paid_orders': 0,
            'paid_gmv': 0.0,
            'pending_shipments': 0,
            'shipped_orders': 0,
            'signed_orders': 0,
            'ship_rate': 0.0,
            'sign_rate': 0.0,
            'avg_ship_hours': 0.0,
            'sync_abnormal_count': 0,
            'stagnant_count': 0,
            'latest_order_at': '',
        })
        bucket['orders'] += 1
        if o.created_at and (not bucket['latest_order_at'] or o.created_at.isoformat() > bucket['latest_order_at']):
            bucket['latest_order_at'] = o.created_at.isoformat()
        if (o.pay_status or '').lower() == 'paid':
            bucket['paid_orders'] += 1
            bucket['paid_gmv'] = round(bucket['paid_gmv'] + _as_float(o.payable_amount), 2)
            if (o.delivery_status or '').lower() not in {'shipped', 'signed'}:
                bucket['pending_shipments'] += 1
        if (o.delivery_status or '').lower() in {'shipped', 'signed'}:
            bucket['shipped_orders'] += 1
        if (o.delivery_status or '').lower() == 'signed':
            bucket['signed_orders'] += 1
        if o.paid_at and o.shipped_at and o.shipped_at >= o.paid_at:
            ship_hours_by_supplier[code].append((o.shipped_at - o.paid_at).total_seconds() / 3600.0)
        for shipment in shipments_by_order.get(o.id, []):
            if _is_sync_abnormal(shipment):
                bucket['sync_abnormal_count'] += 1
            if _is_stagnant(shipment):
                bucket['stagnant_count'] += 1
    rows = []
    for code, bucket in buckets.items():
        paid_orders = bucket['paid_orders']
        shipped_orders = bucket['shipped_orders']
        bucket['ship_rate'] = round((shipped_orders / paid_orders * 100.0), 1) if paid_orders else 0.0
        bucket['sign_rate'] = round((bucket['signed_orders'] / shipped_orders * 100.0), 1) if shipped_orders else 0.0
        hours = ship_hours_by_supplier.get(code, [])
        bucket['avg_ship_hours'] = round(sum(hours) / len(hours), 2) if hours else 0.0
        rows.append(bucket)
    reverse = sort_by not in {'avg_ship_hours', 'supplier_code', 'supplier_name'}
    rows.sort(key=lambda x: x.get(sort_by) or 0, reverse=reverse)
    return {'days': days, 'supplier_code': supplier_code or '', 'rows': rows}


def get_trend(db: Session, days: int = 30, supplier_code: str | None = None) -> dict[str, Any]:
    start, end, _, _ = _period(days)
    orders = _filter_orders(db, start, end, supplier_code).all()
    order_ids = [o.id for o in orders]
    shipments = _shipment_rows(db, order_ids)
    rows = []
    by_day = {d.strftime('%Y-%m-%d'): {'date': d.strftime('%Y-%m-%d'), 'created_orders': 0, 'paid_orders': 0, 'paid_gmv': 0.0, 'shipped_orders': 0, 'signed_orders': 0} for d in _date_range(start, end)}
    for o in orders:
        if o.created_at:
            by_day[o.created_at.strftime('%Y-%m-%d')]['created_orders'] += 1
        if o.paid_at and start <= o.paid_at <= end:
            key = o.paid_at.strftime('%Y-%m-%d')
            by_day[key]['paid_orders'] += 1
            by_day[key]['paid_gmv'] = round(by_day[key]['paid_gmv'] + _as_float(o.payable_amount), 2)
        if o.shipped_at and start <= o.shipped_at <= end:
            by_day[o.shipped_at.strftime('%Y-%m-%d')]['shipped_orders'] += 1
    for s in shipments:
        if s.signed_at and start <= s.signed_at <= end:
            key = s.signed_at.strftime('%Y-%m-%d')
            if key in by_day:
                by_day[key]['signed_orders'] += 1
    rows.extend(by_day.values())
    return {'days': days, 'supplier_code': supplier_code or '', 'rows': rows}


def get_category_supplier_board(db: Session, days: int = 30, supplier_code: str | None = None) -> dict[str, Any]:
    start, end, _, _ = _period(days)
    orders = _filter_orders(db, start, end, supplier_code).all()
    if not orders:
        return {'days': days, 'supplier_code': supplier_code or '', 'rows': []}
    order_map = {o.id: o for o in orders}
    items = db.query(OrderItem).filter(OrderItem.order_id.in_(list(order_map.keys()))).all()
    product_ids = {i.product_id for i in items if i.product_id}
    products = {p.id: p for p in db.query(Product).filter(Product.id.in_(product_ids)).all()} if product_ids else {}
    category_ids = {p.category_id for p in products.values() if p.category_id}
    categories = {c.id: c.name for c in db.query(ProductCategory).filter(ProductCategory.id.in_(category_ids)).all()} if category_ids else {}
    buckets: dict[tuple[str, str], dict[str, Any]] = {}
    for item in items:
        order = order_map.get(item.order_id)
        if not order:
            continue
        supplier = (order.supplier_code or '未分配').strip() or '未分配'
        product = products.get(item.product_id)
        category_name = categories.get(product.category_id, '未分类') if product else '未知商品'
        key = (category_name, supplier)
        bucket = buckets.setdefault(key, {
            'category_name': category_name,
            'supplier_code': supplier,
            'orders': set(),
            'qty': 0,
            'gmv': 0.0,
            'products': set(),
        })
        bucket['orders'].add(order.id)
        bucket['qty'] += int(item.qty or 0)
        bucket['gmv'] = round(bucket['gmv'] + _as_float(item.subtotal), 2)
        if item.product_id:
            bucket['products'].add(item.product_id)
    rows = []
    for (_, _), bucket in buckets.items():
        rows.append({
            'category_name': bucket['category_name'],
            'supplier_code': bucket['supplier_code'],
            'order_count': len(bucket['orders']),
            'qty': bucket['qty'],
            'gmv': bucket['gmv'],
            'product_count': len(bucket['products']),
        })
    rows.sort(key=lambda x: (x['gmv'], x['qty']), reverse=True)
    return {'days': days, 'supplier_code': supplier_code or '', 'rows': rows}


def get_product_ranking(db: Session, days: int = 30, supplier_code: str | None = None, limit: int = 20) -> dict[str, Any]:
    start, end, _, _ = _period(days)
    orders = _filter_orders(db, start, end, supplier_code).all()
    if not orders:
        return {'days': days, 'supplier_code': supplier_code or '', 'rows': []}
    order_map = {o.id: o for o in orders}
    items = db.query(OrderItem).filter(OrderItem.order_id.in_(list(order_map.keys()))).all()
    product_ids = {i.product_id for i in items if i.product_id}
    products = {p.id: p for p in db.query(Product).filter(Product.id.in_(product_ids)).all()} if product_ids else {}
    category_ids = {p.category_id for p in products.values() if p.category_id}
    categories = {c.id: c.name for c in db.query(ProductCategory).filter(ProductCategory.id.in_(category_ids)).all()} if category_ids else {}
    buckets: dict[str, dict[str, Any]] = {}
    for item in items:
        order = order_map.get(item.order_id)
        if not order:
            continue
        product = products.get(item.product_id)
        key = str(item.product_id or item.sku_code or item.product_name)
        bucket = buckets.setdefault(key, {
            'product_id': item.product_id or 0,
            'product_name': item.product_name or (product.name if product else '未知商品'),
            'sku_code': item.sku_code or (product.sku_code if product else ''),
            'category_name': categories.get(product.category_id, '未分类') if product else '未分类',
            'supplier_code': (order.supplier_code or '未分配').strip() or '未分配',
            'orders': set(),
            'qty': 0,
            'gmv': 0.0,
            'paid_orders': 0,
            'signed_orders': 0,
        })
        bucket['orders'].add(order.id)
        bucket['qty'] += int(item.qty or 0)
        bucket['gmv'] = round(bucket['gmv'] + _as_float(item.subtotal), 2)
        if (order.pay_status or '').lower() == 'paid':
            bucket['paid_orders'] += 1
        if (order.delivery_status or '').lower() == 'signed':
            bucket['signed_orders'] += 1
    rows = []
    for bucket in buckets.values():
        rows.append({
            'product_id': bucket['product_id'],
            'product_name': bucket['product_name'],
            'sku_code': bucket['sku_code'],
            'category_name': bucket['category_name'],
            'supplier_code': bucket['supplier_code'],
            'order_count': len(bucket['orders']),
            'qty': bucket['qty'],
            'gmv': bucket['gmv'],
            'paid_orders': bucket['paid_orders'],
            'signed_orders': bucket['signed_orders'],
        })
    rows.sort(key=lambda x: (x['gmv'], x['qty']), reverse=True)
    return {'days': days, 'supplier_code': supplier_code or '', 'rows': rows[: max(1, min(limit, 200))]}


def get_funnel(db: Session, days: int = 30, supplier_code: str | None = None) -> dict[str, Any]:
    start, end, _, _ = _period(days)
    orders = _filter_orders(db, start, end, supplier_code).all()
    created = len(orders)
    pending_payment = sum(1 for o in orders if (o.pay_status or '').lower() != 'paid')
    paid = sum(1 for o in orders if (o.pay_status or '').lower() == 'paid')
    waiting_ship = sum(1 for o in orders if (o.pay_status or '').lower() == 'paid' and (o.delivery_status or '').lower() not in {'shipped', 'signed'})
    shipped = sum(1 for o in orders if (o.delivery_status or '').lower() in {'shipped', 'signed'})
    signed = sum(1 for o in orders if (o.delivery_status or '').lower() == 'signed')
    return {
        'days': days,
        'supplier_code': supplier_code or '',
        'rows': [
            {'stage': '创建订单', 'count': created, 'rate_vs_prev': 100.0},
            {'stage': '待支付', 'count': pending_payment, 'rate_vs_prev': round((pending_payment / created * 100.0), 1) if created else 0.0},
            {'stage': '已支付', 'count': paid, 'rate_vs_prev': round((paid / created * 100.0), 1) if created else 0.0},
            {'stage': '待发货', 'count': waiting_ship, 'rate_vs_prev': round((waiting_ship / paid * 100.0), 1) if paid else 0.0},
            {'stage': '已发货', 'count': shipped, 'rate_vs_prev': round((shipped / paid * 100.0), 1) if paid else 0.0},
            {'stage': '已签收', 'count': signed, 'rate_vs_prev': round((signed / shipped * 100.0), 1) if shipped else 0.0},
        ],
    }


def get_alerts_trend(db: Session, days: int = 30, supplier_code: str | None = None) -> dict[str, Any]:
    start, end, _, _ = _period(days)
    q = db.query(Shipment, Order).join(Order, Order.id == Shipment.order_id)
    if supplier_code:
        q = q.filter(Order.supplier_code == supplier_code)
    q = q.filter(
        ((Shipment.last_sync_at >= start) & (Shipment.last_sync_at <= end)) |
        ((Shipment.updated_at >= start) & (Shipment.updated_at <= end)) |
        ((Shipment.created_at >= start) & (Shipment.created_at <= end))
    )
    pairs = q.all()
    by_day = {d.strftime('%Y-%m-%d'): {'date': d.strftime('%Y-%m-%d'), 'sync_abnormal': 0, 'stagnant': 0, 'exceptional': 0, 'tracked_shipments': 0} for d in _date_range(start, end)}
    for shipment, _order in pairs:
        dt = shipment.last_sync_at or shipment.updated_at or shipment.created_at
        if not dt:
            continue
        key = dt.strftime('%Y-%m-%d')
        if key not in by_day:
            continue
        by_day[key]['tracked_shipments'] += 1
        if _is_sync_abnormal(shipment):
            by_day[key]['sync_abnormal'] += 1
        if _is_stagnant(shipment):
            by_day[key]['stagnant'] += 1
        if _is_exceptional(shipment):
            by_day[key]['exceptional'] += 1
    return {'days': days, 'supplier_code': supplier_code or '', 'rows': list(by_day.values())}


def _sheet_from_rows(title: str, headers: list[str], rows: list[list[Any]]) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    fill = PatternFill('solid', fgColor='1F4E78')
    for cell in ws[1]:
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'
    widths = [max(len(str(c.value or '')) for c in col[: min(len(col), 50)]) + 4 for col in ws.columns]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = min(max(width, 10), 28)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_supplier_board_xlsx(db: Session, days: int = 30, supplier_code: str | None = None) -> tuple[str, BytesIO]:
    data = get_supplier_board(db, days=days, supplier_code=supplier_code)
    rows = [[
        r['supplier_code'], r['supplier_name'], r['mapped_products'], r['orders'], r['paid_orders'], r['paid_gmv'],
        r['pending_shipments'], r['shipped_orders'], r['signed_orders'], r['ship_rate'], r['sign_rate'],
        r['avg_ship_hours'], r['sync_abnormal_count'], r['stagnant_count'], r['latest_order_at']
    ] for r in data['rows']]
    buf = _sheet_from_rows('供应链排行', ['供应链编码', '供应链名称', '映射商品', '订单数', '已支付', '支付GMV', '待发货', '已发货', '已签收', '发货率%', '签收率%', '平均发货时效(h)', '同步异常', '停滞物流', '最近订单时间'], rows)
    filename = f"data_center_supplier_board_{days}d.xlsx"
    return filename, buf


def export_product_ranking_xlsx(db: Session, days: int = 30, supplier_code: str | None = None, limit: int = 50) -> tuple[str, BytesIO]:
    data = get_product_ranking(db, days=days, supplier_code=supplier_code, limit=limit)
    rows = [[
        r['product_id'], r['product_name'], r['sku_code'], r['category_name'], r['supplier_code'], r['order_count'],
        r['qty'], r['gmv'], r['paid_orders'], r['signed_orders']
    ] for r in data['rows']]
    buf = _sheet_from_rows('商品销量榜', ['商品ID', '商品名称', 'SKU', '分类', '供应链', '订单数', '销量', 'GMV', '已支付订单', '已签收订单'], rows)
    filename = f"data_center_product_ranking_{days}d.xlsx"
    return filename, buf
